# !/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Author : 陈坤泽
# @Email  : 877362867@qq.com
# @Date   : 2022/4

""" 惟海法师网课考勤工具

为了使用便捷明了，程序开发上
1、使用中文命名
2、牺牲了一定的工程、扩展灵活性，用了尽可能多的自动推导，而不是参数配置
"""
from pyxllib.prog.pupil import check_install_package

check_install_package('fire')  # 自动安装依赖包

from collections import Counter, defaultdict
from datetime import date, timedelta
import datetime
import math
import os
import re
import time
from io import StringIO
import csv

import fire
import pandas as pd
from tqdm import tqdm
import requests
# import requests_cache

from pyxllib.text.pupil import chinese2digits, grp_chinese_char
from pyxllib.file.xlsxlib import openpyxl
from pyxllib.file.specialist import XlPath, get_encoding
from pyxllib.prog.pupil import run_once
from pyxllib.prog.specialist import parse_datetime, browser, TicToc, dprint

from pyxllib.cv.rgbfmt import RgbFormatter
from pyxllib.data.sqlite import Connection
from pyxllib.ext.seleniumlib import XlChrome


class Xiaoetong:
    """ 写一个觉观的api
    """

    def __init__(self):
        self.token = ''

    def login(self, app_id, client_id, secret_key):
        """ 登录，获取token
        """
        # 启用缓存
        # requests_cache.install_cache('access_token_cache', expire_after=None)  # 设置缓存过期时间xx（单位：秒）
        # 接口地址
        url = "https://api.xiaoe-tech.com/token"
        params = {
            "app_id": app_id,
            "client_id": client_id,
            "secret_key": secret_key,
            "grant_type": "client_credential"
        }
        response = requests.get(url, params=params)
        if response.status_code == 200:
            result = response.json()
            if result['code'] == 0:
                # access_token 是小鹅通开放api的全局唯一接口调用凭据，店铺调用各接口时都需使用 access_token ,开发者需要进行妥善保管；
                self.token = result['data']['access_token']
            else:
                raise Exception("Error getting access token: {}".format(result['msg']))
        else:
            raise Exception("HTTP request failed with status code {}".format(response.status_code))
        # return result['data']['access_token']

    def get_alive_user_list(self, resource_id, page_size=100):
        """ 获取直播间用户
        """
        # 1 获取总页数
        url = "https://api.xiaoe-tech.com/xe.alive.user.list/1.0.0"  # 接口地址【路径：API列表 -> 直播管理 -> 获取直播间用户列表】
        data_1 = {
            "access_token": self.token,
            "resource_id": resource_id,
            "page": 1,
            "page_size": page_size
        }
        response_1 = requests.post(url, data=data_1)
        result_1 = response_1.json()
        page = math.ceil(result_1['data']['total'] / page_size)  # 页数

        # 2 获取直播间用户数据
        lst = result_1['data']['list']
        for i in range(1, page):  # 为什么从1开始，因为第一页的数据上面已经获取到了，这里没必要从新获取一次
            data = {
                "access_token": self.token,
                "resource_id": resource_id,
                "page": i + 1,
                "page_size": page_size
            }
            response = requests.post(url, data=data)
            result = response.json()
            data_1 = result['data']['list']
            lst += data_1
            # lst.extend(data_1)
        return lst

    def get_elock_actor(self, activity_id, page_size=100):
        """ 获取打卡参与用户
        """
        # 获取总页数
        url = "https://api.xiaoe-tech.com/xe.elock.actor/1.0.0"  # 接口地址【路径：API列表 -> 打卡管理 -> 获取打卡参与用户】
        data_1 = {
            "access_token": self.token,
            "activity_id": activity_id,
            "page_index": 1,
            "page_size": page_size
        }
        response_1 = requests.post(url, data=data_1)
        result_1 = response_1.json()
        page = math.ceil(result_1['data']['count'] / page_size)  # 页数
        # 获取打卡用户数据
        lst = result_1['data']['list']
        for i in range(1, page):  # 为什么从1开始，因为第一页的数据上面已经获取到了，这里没必要从新获取一次
            data = {
                "access_token": self.token,
                "activity_id": activity_id,
                "page_index": i + 1,
                "page_size": page_size
            }
            response = requests.post(url, data=data)
            result = response.json()
            data_1 = result['data']['list']
            lst += data_1
            # lst.extend(data_1)
        return lst


class 网课考勤:
    def __init__(self, today=None, *, 表格路径=None):
        self.返款标题 = ''
        self.表格路径 = 表格路径 or r'考勤.xlsx'
        self.在线表格 = 'https://docs.qq.com/sheet/DUlF1UnRackJ2Vm5U'  # 生成日报用
        self.开课日期 = '2022-01-08'
        self.视频返款 = [20, 15, 10, 5, 0, 0]  # 直播(当堂)/第1天（当天）/第2天/第3天/第4天/第5天，完成观看的依次返款额。
        self.打卡返款 = {5: 100, 10: 150, 15: 200}  # 打卡满5/10/15次的返款额
        self.课程链接 = []
        self._init(today)

        self.driver = None  # 浏览器脚本

    def __1_基础数据处理(self):
        pass

    def _init(self, today=None):
        """ 可以手动指定today，适合某些场景下的调试 """
        self.课次数 = len(self.课程链接) - 1  # 课次一般都是固定21课，不用改
        self.达标时长 = 30  # 单位:分钟。在线听课的时长达到此标准以上，才计算为完成一节课的学习
        # 一般是3天，因为回放第1、2、3天都会数额，第4天就没数额了，所以第4天返款
        self.回放返款天数 = sum((x > 0) for x in self.视频返款[1:])

        self.root = XlPath(self.表格路径).parent
        self.wb = openpyxl.load_workbook(self.表格路径)
        self.ws = self.wb['考勤表']
        if today:
            dt = parse_datetime(today)
            self.today = date(dt.year, dt.month, dt.day)
        else:
            self.today = date.today()
        self.开课日期 = date.fromisoformat(self.开课日期)
        # self.觉观禅课 = (self.开课日期.day == 1)
        self.觉观禅课 = '觉观' in self.返款标题
        self.当天课次 = (self.today - self.开课日期).days + 1  # 这是用于逻辑运算的，可能超过实际课次数
        self.当天课次2 = min(self.当天课次, self.课次数)
        self.结束课次 = self.当天课次 - len(self.视频返款) + 1  # 这是用于逻辑运算的，可能有负值
        self.结束课次2 = max(0, self.结束课次)
        try:
            self.用户列表 = pd.read_csv(self.get_file('数据表/用户列表导出*.csv'), dtype={16: str})
        except:
            self.用户列表 = None
        # 用户列表 = 用户列表[用户列表['账号状态'] == '正常']

        self.考勤表出现次数 = Counter()
        for f in self.root.glob('数据表/**/*直播观看详情*.csv'):
            df = pd.read_csv(f, skiprows=1)
            self.考勤表出现次数 += Counter([x.strip() for x in df['用户ID']])
        for f in self.root.glob('数据表/**/*直播用户列表*.csv'):
            df = pd.read_csv(f)
            self.考勤表出现次数 += Counter([x.strip() for x in df['用户ID']])

        self.异常data = {}

        if hasattr(self, '异常处理'):
            self.异常处理()

    def get_file(self, p):
        # 返回最后一个匹配的文件
        return list(self.root.glob(p))[-1]

    def 查找用户(self, 昵称='', 手机号='', *, debug=False):
        """
        :param 昵称: 昵称可以输入一个列表，会在目标昵称、真实姓名同时做检索
        :param 手机号: 手机也可以输入一个列表，因为有些人可能报名时手机号填错，可以增加一些匹配规则
        :return:
        """

        # 1 统一输入参数格式
        def tolist(x):
            if x and not isinstance(x, list):
                x = [x]
            x = [str(a) for a in x if a]
            return x

        def check_telphone(v, refs):
            if not isinstance(v, str):
                try:
                    v = int(v)
                except:
                    pass
                v = str(v)
            for a in refs:
                if a in v:
                    return True

        昵称 = tolist(昵称)
        手机号 = tolist(手机号)
        手机号 = [f'{x}' for x in 手机号 if (x and x != 'None')]

        # 2 查找所有可能匹配的项目
        ls = []
        for idx, x in self.用户列表.iterrows():
            logo = 0
            if '真实姓名' in x:
                x['姓名'] = x['真实姓名']

            if 昵称:
                if x['昵称'] in 昵称 or x['姓名'] in 昵称:
                    logo += 1
            if 手机号:
                if check_telphone(x.get('账户绑定手机号'), 手机号) or check_telphone(x.get('最近采集的手机号'), 手机号):
                    logo += 2
            if logo:
                ls.append([x, logo])

        ls.sort(key=lambda x: x[1], reverse=True)
        ls = [x[0] for x in ls]

        if debug:
            print('\n'.join(map(self.用户信息摘要, ls)))
            return

        return ls

    def 用户信息摘要(self, x):
        """ 输入series类型的一个条目，输出其摘要信息 """
        ls = [f'{k}={v}' for k, v in x.items() if (not isinstance(v, float) or not math.isnan(v))]
        if x['用户ID'] in self.考勤表出现次数:
            ls.append('考勤' + str(self.考勤表出现次数[x['用户ID']]) + '次')
        return ', '.join(ls)

    def 匹配用户ID(self, sheet_name='报名表'):

        def try2int(x):
            try:
                return int(x)
            except:
                if isinstance(x, float) and math.isnan(x):
                    return ''
                return x

        def todict(row):
            """ 将表格第i行的数据提取为字典格式
            """
            msg = {}
            for k, x in enumerate(['真实姓名', '微信昵称', '手机号', '错误手机号']):
                msg[x] = ws.cell2(row, x).value
            return msg

        ws = self.wb[sheet_name]
        for row in tqdm(list(ws.iterrows('真实姓名')), desc='匹配进度'):
            # todo 这里要考虑相似度整体权重的，而不是几个属性的优先级
            x = todict(row)
            x['手机号'] = str(x['手机号']).lstrip('`')
            x['错误手机号'] = str(x['错误手机号']).lstrip('`')

            待查手机号 = [x['手机号']]
            t = try2int(x['错误手机号'])
            if t:
                待查手机号.append(t)
            ls = self.查找用户([x['真实姓名'], x['微信昵称']], 待查手机号)
            摘要ls = list(map(self.用户信息摘要, ls))

            用户ID = ''
            if len(ls) == 1:
                # 只有一个关联的直接匹配上，否则填空
                用户ID = ls[0]['用户ID']
            else:
                # 只有一条有考勤记录时
                flags = [('考勤' in text) for text in 摘要ls]
                if sum(flags) == 1:
                    idx = flags.index(1)
                    用户ID = ls[idx]['用户ID']
            if 用户ID == '':
                flags = [('账号状态=正常' in text) for text in 摘要ls]
                if sum(flags) == 1:
                    idx = flags.index(1)
                    用户ID = ls[idx]['用户ID']
            # 如果用户ID还是空的，则用手机号能匹配上的那条
            if 用户ID == '':
                for i, text in enumerate(摘要ls):
                    if str(x['手机号']) in text:
                        用户ID = ls[i]['用户ID']
                        break

            ws.cell2(row, '用户ID', 用户ID)
            ws.cell2(row, '参考信息', '\n'.join(摘要ls))

            row += 1
        self.wb.save(self.表格路径)

    def 获取每日统计表(self):
        ls = []
        columns = ['课次', '用户ID', '观看日期', '在线时长(分钟)']
        for f in self.root.glob('数据表/*.csv'):
            if m := re.match(r'(\d{4}\-\d{2}\-\d{2}).+?课.*?(\d+).+?直播观看详情', f.stem):
                stat_day, 课次 = date.fromisoformat(m.group(1)), int(m.group(2))
                skiprows = 1
            elif m := re.search(r'届(?:念住|觉观).+?(\d+).+?直播用户列表.+?(\d{4}\-\d{2}\-\d{2})', f.stem):
                stat_day, 课次 = date.fromisoformat(m.group(2)), int(m.group(1))
                skiprows = 0
            elif m := re.search(r'第(\d+)堂.+?届(?:念住|觉观).+?直播用户列表.+?(\d{4}\-\d{2}\-\d{2})', f.stem):
                stat_day, 课次 = date.fromisoformat(m.group(2)), int(m.group(1))
                skiprows = 0
            elif m := re.search(r'本体音艺网课-(\d+).+?直播用户列表.+?(\d{4}\-\d{2}\-\d{2})', f.stem):
                stat_day, 课次 = date.fromisoformat(m.group(2)), int(m.group(1))
                skiprows = 0
            else:
                continue

            if stat_day > self.today:  # 超过self.today的数据不记录
                continue
            观看日期 = (stat_day - self.开课日期).days - 课次 + 1
            df = pd.read_csv(f, skiprows=skiprows)
            for idx, r in df.iterrows():
                k = '累计观看时长(秒)' if 观看日期 else '直播观看时长(秒)'
                ls.append([课次, r['用户ID'].strip(), 观看日期, int(int(r[k]) / 60)])

        # ext: 异常数据修正，trick
        for k, v in self.异常data.items():
            课次, 用户ID = k
            # 观看日期和在线时长可以乱填，主要是把这个用户课次记录下，后面遍历才会处理到
            ls.append([int(re.search(r'\d+', 课次).group()), 用户ID, 0, 0])

        df = pd.DataFrame.from_records(ls, columns=columns)
        return df

    def 获取考勤表中出现的所有用户名(self):
        df = self.获取每日统计表()
        print(*set(df['用户ID']), sep='\n')

    def 修订(self, 学号, 课次, 完成标记, *args):
        try:
            user_id = self.ws.cell2({'学号': 学号}, '用户ID').value
        except ValueError:
            return
        if isinstance(课次, int):
            课次 = f'第{课次:02}课'
        self.异常data[课次, user_id] = 完成标记

    def 更新统计表(self, df):
        # 1 辅助函数
        ws = self.ws
        id2row = {}
        cel = ws.findcel('用户ID').down()
        while cel.value:
            id2row[cel.value] = cel.row
            cel = cel.down()

        col = ws.findcol('第01课') - 1

        def 完成标记(items):
            """ 输入某个用户某个课次每天的累计观看时长 """
            x = {t['观看日期']: t['在线时长(分钟)'] for idx, t in items.iterrows()}
            for k in sorted(x.keys()):
                if x[k] >= self.达标时长:
                    return f'第{k}天回放' if k else '完成当堂学习'
            return f'不足{self.达标时长}分钟'

        def write(课次, 用户ID, value):
            # 优先使用修订的标记
            value = self.异常data.get((f'第{课次:02}课', 用户ID), value)
            cel = ws.cell2(id2row[用户ID], f'第{课次:02}课')
            if cel is None:
                return
            cel.value = value
            color = None
            if '完成当堂学习' in value:
                color = RgbFormatter.from_name('鲜绿色')
            elif '回放' in value:
                color = RgbFormatter.from_name('黄色')
                v1 = self.视频返款[0]
                idx = min(int(re.search(r'第(\d+)天', value).group(1)),
                          len(self.视频返款) - 1)
                v2 = self.视频返款[idx]

                if v2:
                    color = color.light((v1 - v2) / v2)  # 根据返款额度自动变浅
                else:  # 如果无返款额度
                    color = RgbFormatter.from_name('灰色')
            elif 课次 <= self.当天课次 - len(self.视频返款) + 1:
                cel.value = '未完成学习'
                color = RgbFormatter.from_name('红色')

            if color:
                cel.fill_color(color)

        # 2 遍历更新考勤情况数据
        for [课次, 用户ID], items in df.groupby(['课次', '用户ID']):
            if 用户ID not in id2row:
                continue
            write(课次, 用户ID, 完成标记(items))

        # 3 处理剩余用户
        for i in range(1, min(self.课次数 + 1, self.结束课次 + 1)):
            for k, v in id2row.items():
                t = ws.cell(v, col + i).value
                if not t or t in ('未开始学习', f'不足{self.达标时长}分钟'):
                    write(i, k, '未完成学习')
        for i in range(max(self.结束课次 + 1, 1), min(self.课次数 + 1, self.当天课次 + 1)):
            for k, v in id2row.items():
                if not ws.cell(v, col + i).value:
                    write(i, k, '未开始学习')

    def 生成今日返款文件(self, msg):
        """ 生成今日返款文件，及缺勤报告

        生成的csv文件共4列：订单号，反馈额，备注说明，重复校验码
        第4列可以设置一个64长度以内的标识来判重，防止重复退款
            因为每个学员每个课次理论上应该只有一次退款操作
            所以我使用"{订单号}_class{课次号}"来生成校验码

        注意，涉及给人看的数据，大部分课次编号都不会加前导0，
        涉及程序、文件名对齐的，则一般都会设前导0
        """
        ls = []  # 返款汇总
        name = []  # 生成的csv文件名

        ws = self.ws
        _col = ws.findcol('第01课') - 1

        # 1 初期要详细提示
        if self.当天课次 < 4:
            msg.append('第1次建议大家都查看下完整考勤表，记一下自己的学号，方便以后核对考勤数据。'
                       '以后群里统一以学号的方式汇报异常的考勤数据，未列出则都是"正常完成当堂学习"的学员。')
            msg.append('如果发现自己考勤数据不对，可以群里、私聊反馈给我修正。')

        # 2 要过期的课程
        if 0 < self.结束课次 + 1 <= self.课次数:
            msg.append(f'第{self.结束课次 + 1}课回放、打卡第{len(self.视频返款) - 1}天，还未完成学习的同学们请抓紧时间。')

        # 3 回放
        回放课次 = self.当天课次 - self.回放返款天数
        if 0 < 回放课次 <= self.课次数:
            title = f'第{回放课次}课回放'
            msg.append('【' + title + '名单】')
            if 回放课次 == 1:
                msg[-1] += f'（第1课仍有{len(self.视频返款) - self.回放返款天数 - 1}天可以回放）'

            d = {}
            for i in range(1, self.回放返款天数 + 1):
                d[f'第{i}天回放'] = []
            d['未完成学习'] = []

            col = _col + 回放课次
            for r in ws.iterrows('用户ID'):
                v = ws.cell(r, col).value
                if v and '回放' in v:
                    t = int(re.search(r'第(\d+)天', v).group(1))
                    if self.视频返款[t]:
                        订单号 = ws.cell2(r, '交易订单号').value
                        if 订单号:
                            cols = [订单号, self.视频返款[t],
                                    f'{self.返款标题}第{回放课次}课第{t}天完成回放',
                                    f'{订单号}_class{回放课次:02}']  # 防止重复返款的校验码
                            ls.append(','.join(map(str, cols)))
                    else:
                        v = '未完成学习'
                elif v in (f'不足{self.达标时长}分钟', '未开始学习'):
                    v = '未完成学习'
                if v and v != '完成当堂学习':
                    d[v].append(ws.cell2(r, '学号').value)

            is_empty = True
            for k, v in d.items():
                if v:
                    m = re.search(r'\d+', k)
                    t = int(m.group()) if m else 5
                    msg[-1] += f'\n{k}(返{self.视频返款[t]}元): ' + ','.join(map(str, v))
                    is_empty = False
            if is_empty:
                msg[-1] = f'第{回放课次}课因当堂满勤，无回放名单'
            else:
                name.append(f'第{回放课次}课回放')

        # 4 当堂
        if self.当天课次 <= self.课次数:
            title = f'第{self.当天课次}课当堂'
            msg.append('【' + title + '缺勤名单】')
            if self.当天课次 == 1:
                msg[-1] += '（注意只统计早晨直播情况，今日的回放数据明天会更新）'
            # d = {f'不足{self.达标时长}分钟': [], '未开始学习': []}
            d = []

            col = _col + self.当天课次
            for r in ws.iterrows('用户ID'):
                v = ws.cell(r, col).value
                if v == '完成当堂学习':
                    订单号 = ws.cell2(r, '交易订单号').value
                    if 订单号:
                        cols = [订单号, self.视频返款[0],
                                f'{self.返款标题}第{self.当天课次}课完成当堂学习',
                                f'{订单号}_class{self.当天课次:02}']  # 防止重复返款的校验码
                        ls.append(','.join(map(str, cols)))
                else:
                    d.append(ws.cell(r, 1).value)

            if d:
                msg[-1] += '\n' + ','.join(map(str, d))
            else:
                msg[-1] = title + '满勤！！！'
            name.append(f'第{self.当天课次}课当堂')

        # 5 第22课的处理
        if self.觉观禅课 and self.当天课次 == 23:
            name.append('第22课')
            for r in ws.iterrows('用户ID'):
                v = ws.cell2(r, '交易订单号').value
                if v:
                    cols = [v, self.视频返款[0],
                            f'{self.返款标题}第{self.当天课次 - 1}课',
                            f'{v}_class22']  # 防止重复返款的校验码
                    ls.append(','.join(map(str, cols)))

        # 6 返款文件
        if ls:
            ls = [x for x in ls if ('订单号' not in x and 'None' not in x)]
            (self.root / (f'第{self.当天课次:02}天 ' + '+'.join(name) + '返款.csv')).write_text('\n'.join(ls))

        # 7 提示信息
        if name:
            msg.append('+'.join(name) + '促学金已返款，同学们请查收。')
        if self.觉观禅课 and self.当天课次 == 22:
            msg.append('今晚第22课答疑不考勤，明天统一返款。')
        return msg

    def 计算打卡返款(self, 打卡次数):
        from bisect import bisect_right

        ks = [0] + list(self.打卡返款.keys())
        ks.sort()

        def find_le(a, x):
            'Find rightmost value less than or equal to x'
            i = bisect_right(a, x)
            if i:
                return a[i - 1]
            raise ValueError

        k = find_le(ks, 打卡次数)
        return self.打卡返款.get(k, 0)

    def 打卡统计(self, msg):
        """ 不知道平台有每个课程的打卡次数，自己用另外一个途径暴力搞的，220226周六15:59 现在应该没用了"""
        from openpyxl.styles import PatternFill

        # 1 至少两个目录才统计
        dirs = list((self.root / '数据表').glob_dirs('用户学习统计*'))
        if len(dirs) < 2:
            return

        # 2 读取数据
        def merge_data(d):
            """ 输入目录，汇总目录下所有xlsx文件的打卡数据

            导出的打卡数据文件尽量小些，不然这个操作速度有点慢。。。
            """
            data = {}
            for f in d.glob('*.xlsx'):
                if f.stem.startswith('~$'):
                    continue
                df = pd.read_excel(f, header=1)
                for idx, x in df.iterrows():
                    user_id = x['用户ID\t']
                    if user_id == user_id:  # nan数值的特性，这个条件不成立
                        value = x['提交打卡\t']
                        data[user_id.strip()] = int(value) if value == value else 0
                else:
                    continue

            return data

        d1, d2 = dirs[0], dirs[-1]
        data1 = merge_data(d1)
        data2 = merge_data(d2)

        # 3 d1数据标准化
        # 最理想的是d1恰好是开课前一天的统计数据，那么d2和d1直接相减即可
        # 如果d1提前了，无法判断具体打卡日期，所以还是直接相减，只能都算上
        # 如果d1延后了，那么要对d1的数据进行预处理，要减掉对应天数，但最低为0，不能为负数
        def parse_date(name):
            """ 输入文件名，获得其日期 """
            m = re.search(r'(\d{2})(\d{2})(\d{2})', name)
            year, month, day = m.groups()
            year = '20' + year
            return date(int(year), int(month), int(day))

        delta = (parse_date(d1.stem) - self.开课日期).days + 1
        if delta > 0:
            data1 = {k: max(0, v - delta) for k, v in data1.items()}

        # 4 计算新的打卡次数
        # 打卡数是有可能超过21次的
        data3 = {k: v - data1.get(k, 0) for k, v in data2.items()}

        # 5 将打卡次数写入表格
        ls = []  # 返款汇总
        ws = self.ws
        color0 = [(255, 0, 0), (255, 255, 128), (255, 255, 0), (0, 255, 0)]
        for i in ws.iterrows('用户ID'):
            c1 = ws.cell2(i, ['打卡返款', '打卡数'])
            v = c1.value = data3.get(ws.cell2(i, '用户ID').value, 0)
            v //= 5
            if v:
                money = self.打卡返款[min(v - 1, 2)]
                订单号 = ws.cell2(i, '交易订单号').value
                cols = [订单号, money, f'{self.返款标题}返学修日志促学金', f'{订单号}_journal']
                if cols[0]:
                    ls.append(','.join(map(str, cols)))
            else:
                money = 0

            c2 = ws.cell2(i, ['打卡返款', '返款'])
            c2.value = money

            color = RgbFormatter(*color0[min(3, v)])
            c1.fill = PatternFill(fgColor=color.hex[-6:], fill_type="solid")
            c2.fill = PatternFill(fgColor=color.hex[-6:], fill_type="solid")

        # 并计算当前的返款额
        desc = '/'.join(map(str, self.打卡返款))
        if self.当天课次 == 25:
            msg.append('已生成截止目前的打卡数据，同学们可以预先核对下，明天最后更新打卡数据后返款。'
                       f'打卡达到"5/10/15"次，依次返回"{desc}"元。'
                       '注：因技术原因，打卡数据无法精确计算，统计遵循宁可多算但无漏算的原则，所以部分同学打卡数会超过21次。')
        elif self.当天课次 == self.课次数 + len(self.视频返款) - 1:
            # 生成打卡返款
            msg.append('已完成学修日志（打卡）促学金的返款。'
                       f'打卡达到"5/10/15"次，依次返回"{desc}"元。')
            (self.root / '学修日志返款.csv').write_text('\n'.join(ls))

    def 打卡统计2(self, msg):
        from openpyxl.styles import PatternFill

        # 1 有打卡统计表才计算
        files = list((self.root / '数据表').glob_files('*打卡*.csv'))
        if len(files) < 1:
            return

        # 2 读取数据
        # df = pd.read_csv(files[-1])
        # data = {}
        # for idx, row in df.iterrows():
        #     data[row['用户id']] = row['打卡次数']

        df = None
        try:
            df = pd.read_excel(files[-1])  # 220804周四08:28，小鹅通更新了模板
        except ValueError:
            pass

        if df is None:
            try:
                df = pd.read_csv(files[-1])  # 221005周三09:19，小鹅通又双叒更新了
            except UnicodeDecodeError:
                pass

        if df is None:
            try:
                df = pd.read_csv(files[-1], encoding="ANSI")  # 240226周一11:21，
            except UnicodeDecodeError:
                pass

        if df is None:
            raise ValueError

        df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        df.columns = df.columns.map(lambda x: x.strip() if isinstance(x, str) else x)

        # data = Counter([x for x in df['用户id']])
        try:
            data = {row['用户id']: row['打卡次数'] for _, row in df.iterrows()}
        except KeyError:  # 230202周四19:49，另一种实际是xlsx格式，然后再转出csv的情况
            data = Counter([row['user_id'] for _, row in df.iterrows()])

        # 3 将打卡次数写入表格
        ls = []  # 返款汇总
        ws = self.ws
        最高返款额 = max(self.打卡返款.values())
        for i in ws.iterrows('用户ID'):
            c1 = ws.cell2(i, ['打卡返款', '打卡数'])
            打卡次数 = c1.value = data.get(ws.cell2(i, '用户ID').value, 0)
            返款额 = self.计算打卡返款(打卡次数)

            if 返款额 > 0:
                订单号 = ws.cell2(i, '交易订单号').value
                cols = [订单号, 返款额, f'{self.返款标题}返学修日志促学金', f'{订单号}_journal']
                ls.append(','.join(map(str, cols)))

            c2 = ws.cell2(i, ['打卡返款', '返款'])
            c2.value = 返款额

            if 返款额 <= 0:
                color = RgbFormatter(255, 0, 0)
            elif 返款额 < 最高返款额:
                color = RgbFormatter(255, 255, 0).light(1 - 返款额 / 最高返款额)
            elif 返款额 == 最高返款额:  # 最高返款额
                color = RgbFormatter(0, 255, 0)

            c1.fill = PatternFill(fgColor=color.hex[-6:], fill_type="solid")
            c2.fill = PatternFill(fgColor=color.hex[-6:], fill_type="solid")

        if ls:
            ls = [x for x in ls if ('无订单号' not in x and 'None' not in x)]

        # 4 生成通知，及返款文件
        desc1 = '/'.join(map(str, self.打卡返款.keys()))
        desc2 = '/'.join(map(str, self.打卡返款.values()))
        if self.结束课次 == self.课次数 - 1:
            msg.append('已生成截止目前的打卡数据，同学们可以预先核对下，明天最后更新打卡数据后返款。'
                       f'打卡达到"{desc1}"次，依次返回"{desc2}"元。')
        elif self.结束课次 == self.课次数:
            # 生成打卡返款
            msg.append('已完成学修日志（打卡）促学金的返款。'
                       f'打卡达到"{desc1}"次，依次返回"{desc2}"元。')
            (self.root / '学修日志返款.csv').write_text('\n'.join(ls))

    def write_返款总计(self):
        ws = self.ws
        name2idx = {k: i for i, k in enumerate(['完成当堂学习', '第1天回放', '第2天回放', '第3天回放', '第4天回放'])}
        for i in ws.iterrows('用户ID'):
            total = 0
            for k in range(1, 22):
                cel = ws.cell2(i, f'第{k:02}课')
                if not cel:
                    continue
                v = cel.value
                if v in name2idx:
                    idx = name2idx[v]
                    if idx and k > self.当天课次 - self.回放返款天数:  # 还没返款的回放金额
                        continue
                    total += self.视频返款[idx]
            v2 = ws.cell2(i, ['打卡返款', '返款']).value
            if isinstance(v2, int):
                total += v2
            if self.觉观禅课 and self.当天课次 > 22:
                total += self.视频返款[0]
            ws.cell2(i, ['已返款', '总计'], total)

    def 考勤日报(self, debug=False, journal=False):
        msg = []  # 报告内容

        # 1 日常返款
        df = self.获取每日统计表()
        self.更新统计表(df)
        self.生成今日返款文件(msg)

        # 2 打卡返款
        if journal or self.结束课次 == self.课次数:
            self.打卡统计2(msg)
        self.write_返款总计()

        # 3 结课
        if self.当天课次 == (self.课次数 + len(self.视频返款) - 1):
            msg.append('考勤返款工作已全部完成，若对自己促学金还有疑问的近日可以再跟我反馈。')

        # 4 输出日报
        if len(msg) > 1:
            msg = '\n'.join([f'{i}、{x}' for i, x in enumerate(msg, start=1)])  # 编号
        else:
            msg = msg[0]
        if self.在线表格:
            msg = f'【完整考勤表】{self.在线表格}\n' + msg
        print(msg)

        # 5 展示或保存表格内容
        if debug:
            browser.html(self.ws.to_html())
        else:
            # self.wb.save(self.表格路径)
            self.wb.save(self.root / (XlPath(self.表格路径).stem + '.xlsx'))

    def cmd(self):
        """ 初始化类后不关闭，继续轮询执行功能 """
        cmd = input('>')
        while cmd != 'exit':
            fire.Fire(self, cmd)
            cmd = input('>')

    def __call__(self, *args, **kwargs):
        print('程序测试通过，可正常使用')

    def 表格内容对齐(self, ws_name1, ws_name2, col_name):
        wb = self.wb
        wb.merge_sheets_by_keycol([wb[ws_name1], wb[ws_name2]], col_name)

        # 保存。一般不用重新再加载self.wb、self.ws，因为这里需求本来基本都是要分段重新执行程序的。
        wb.save(self.表格路径)

    def 匹配交易单号(self, value):
        """ 更加智能的一条龙匹配操作

        :param value: 需要输入待检索的金额值
        """
        # 1 读取账单，账单文件可能不唯一，可能有重复可以自动去重
        files = XlPath('数据表').rglob_files('*基本账户*.csv')
        df_list = []
        for f in files:
            df = pd.read_csv(f)
            df_list.append(df)
        df = pd.concat(df_list, ignore_index=True)
        # 按照"资金流水单号"去重
        df = df.drop_duplicates(subset=['资金流水单号'], keep='first')
        df = df[df['收支类型'] == '`收入']
        df.reset_index(drop=True, inplace=True)

        # 2 因为情况比较特殊，这里不调用通用的对齐功能，而是定制化写过
        ws = self.wb['报名表']

        data = ws.iterrows('真实姓名', to_dict=['交易单号'])
        last_row = -1
        for i, row in data:
            last_row = i
            # 在 df['微信支付业务单号'] 找是否有 row['交易单号']
            items = df[df['微信支付业务单号'] == row['交易单号']]
            if items.empty:
                continue

            # 如果匹配，理论上只有一条
            item = items.iloc[0]
            ws.cell2(i, '交易订单号').value = item['业务凭证号'][1:]
            ws.cell2(i, '订单金额').value = item['收支金额(元)']

            # 在df中去掉所有items
            df = df.drop(items.index)

        # 3 匹配完后，还有目标金额的数据要列出来

        # items = df[df['收支金额(元)'] == f'`{value:.2f}']
        # # todo 这个要识别到'交易单号'中间的日期，降序排序
        # for idx, row in items.iterrows():
        #     last_row += 1
        #     ws.cell2(last_row, '交易单号').value = row['微信支付业务单号']
        #     ws.cell2(last_row, '交易订单号').value = row['业务凭证号'][1:]
        #     ws.cell2(last_row, '订单金额').value = row['收支金额(元)']

        # todo：新增测试用：
        # 初始化一个空的 DataFrame，用于存储遍历结果
        result_df = pd.DataFrame(columns=['微信支付业务单号', '业务凭证号', '收支金额'])
        # 遍历符合条件的数据并存储到 result_df 中
        items = df[df['收支金额(元)'] == f'`{value:.2f}']
        for idx, row in items.iterrows():
            # 将 row 中的数据添加到 result_df 中
            result_df = pd.concat([
                result_df,
                pd.DataFrame({
                    '微信支付业务单号': [row['微信支付业务单号']],
                    '业务凭证号': [row['业务凭证号'][1:]],  # 去掉业务凭证号的第一个字符
                    '收支金额': [row['收支金额(元)']]
                })
            ], ignore_index=True)

        # 提取日期部分，并将其转为可排序的格式
        result_df['日期'] = result_df['微信支付业务单号'].str[11:19]  # 提取日期部分，假设日期位于固定位置
        result_df['日期'] = pd.to_datetime(result_df['日期'], format='%Y%m%d')

        # 按日期降序排序
        result_df = result_df.sort_values(by='日期', ascending=False).drop(columns=['日期'])

        # 将排序后的 DataFrame 写入到工作表中
        for idx, row in result_df.iterrows():
            last_row += 1
            ws.cell2(last_row, '交易单号').value = row['微信支付业务单号']
            ws.cell2(last_row, '交易订单号').value = row['业务凭证号']
            ws.cell2(last_row, '订单金额').value = row['收支金额']

        # 4 保存
        self.wb.save(self.表格路径)

    def __2_自动浏览网页(self):
        pass

    def ensure_driver(self):
        if not hasattr(self, 'driver'):
            self.driver = None
        if not self.driver:
            self.driver = XlChrome()
        return self.driver

    def 登录小鹅通(self, name, passwd):
        # 登录小鹅通
        driver = self.ensure_driver()
        driver.get('https://admin.xiaoe-tech.com/t/login#/acount')
        driver.locate('//*[@id="common_template_mounted_el_container"]'
                      '/div/div[1]/div[3]/div/div[4]/div/div[1]/div[1]/div/div[2]/input').send_keys(name)
        driver.locate('//*[@id="common_template_mounted_el_container"]'
                      '/div/div[1]/div[3]/div/div[4]/div/div[1]/div[2]/div/div/input').send_keys(passwd)
        driver.click('//*[@id="common_template_mounted_el_container"]/div/div[1]/div[3]/div/div[4]/div/div[2]')

        # 然后自己手动操作验证码
        # 以及选择"店铺"

    def 下载课次考勤数据(self, 起始课=None, 终止课=None, 文件名前缀=''):
        if 起始课 is None:
            起始课 = max(1, self.结束课次)
        if 终止课 is None:
            终止课 = min(self.当天课次, self.课次数)

        # 1 遍历课程下载表格
        driver = self.driver
        for i in range(起始课, 终止课 + 1):
            print(f'第{i}课', self.课程链接[i])
            driver.get('https://admin.xiaoe-tech.com/t/data_center/index')  # 必须要找个过渡页，不然不会更新课程链接
            driver.get(self.课程链接[i])
            # 不能写'第{i}课'，会有叫'第{i}堂'等其他情况
            driver.locate_text('//*[@id="app"]/div/div/div[1]/div[2]/div[1]/div[2]', f'{i}')  # 检查页面

            driver.click('//*[@id="tab-studentTab"]/span')  # 直播间用户
            driver.click('//*[@id="pane-studentTab"]/div/div[2]/div[2]/form/div[2]/button[2]/span/span')  # 导出列表
            driver.click('//*[@id="data-export-container"]/div/div[2]/div/div[2]/div[2]/button[2]/span/span')  # 导出
            time.sleep(1)

        # 2 下载表格
        # 2.1 等待下载文件生成完毕
        while True:
            driver.get('https://admin.xiaoe-tech.com/t/basic-platform/downloadCenter')
            if '任务撤回' in driver.locate('//*[@id="downloadsCenter"]/div[4]/div/div[4]/div[2]/table/tbody/tr[1]/'
                                           'td[9]/div/div/span').text:
                time.sleep(1)
            else:
                time.sleep(2)  # 不能下载太快，还是要稍微等一会
                break

        # 2.2 下载表格
        files = []
        for i in range(1, 终止课 - 起始课 + 2):
            driver.click(f'//*[@id="downloadsCenter"]/div[4]/div/div[4]/div[2]/table/tbody/tr[{i}]/'
                         f'td[9]/div/div/span[1]')
            file = driver.locate(f'//*[@id="downloadsCenter"]/div[4]/div/div[3]/table/tbody/tr[{i}]/'
                                 f'td[1]/div/div/div').text + '.csv'  # 下载后还会再多一个csv后缀
            files.append(file)
            time.sleep(1)

        # 2.3 移动表格
        time.sleep(1)
        # 这样应该也不是很泛用，有的人浏览器下载可能会放到各种自定义的其他地方
        download_path = XlPath(os.path.join(os.environ['USERPROFILE'], 'Downloads'))
        for file in files:
            src_file = (download_path / file)
            while True:
                if src_file.is_file():
                    src_file.move(self.root / '数据表' / (文件名前缀 + file), if_exists='replace')
                    break
                else:
                    time.sleep(1)


class 网课考勤2(网课考勤):
    def login_xe(self):
        self.xe = Xiaoetong()  # 实例化
        self.xe.login(self.app_id,
                      self.client_id,
                      self.secret_key)  # 获取了token

        return self.xe.token

    # 依据课程链接，获取资源id（与课次）
    def 获取课次与资源id(self):
        课程链接 = self.课程链接[1:]
        ls_resource_id = [""]
        for item in 课程链接:  # 课次
            resource_id = re.search(r"detail\?id=(.+?)\&", item)  # 资源id
            ls_resource_id.append(resource_id.group(1))
        return ls_resource_id

    # 获取直播间用户数据
    def 获取直播间用户数据(self, resource_id, path):
        if path.is_file():
            return
            # 2）获取直播间用户数据：
        lst = self.xe.get_alive_user_list(resource_id)
        fieldnames = ['用户ID', '用户昵称', '备注名', '状态', '直播间停留时长(秒)', '直播间停留时长',
                      '累计观看时长(秒)', '累计观看时长', '直播观看时长(秒)', '直播观看时长', '回放观看时长(秒)',
                      '回放观看时长', '评论次数', '直播间成交金额']
        p = path
        with open(p, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            writer.writeheader()
            for x in lst:
                record = {
                    '用户ID': x['user_id'],
                    '用户昵称': x['wx_nickname'],
                    '备注名': None,
                    '状态': "其它关联权益",
                    '直播间停留时长(秒)': x['his_online_time'],
                    '直播间停留时长': str(timedelta(seconds=x['his_online_time'])),
                    '累计观看时长(秒)': x['his_learn_time'],
                    '累计观看时长': str(timedelta(seconds=x['his_learn_time'])),
                    '直播观看时长(秒)': x['his_learning_time'],
                    '直播观看时长': str(timedelta(seconds=x['his_learning_time'])),
                    '回放观看时长(秒)': x['his_learned_time'],
                    '回放观看时长': str(timedelta(seconds=x['his_learned_time'])),
                    '评论次数': x['comment_num'],
                    '直播间成交金额': x['user_total_price']
                }
                writer.writerow(record)

    def 获取课次列表(self):
        return list(range(max(self.结束课次2, 1), self.当天课次2 + 1))

    def 下载课程(self):
        prfx = self.prfx
        lt = self.获取课次与资源id()
        for i in tqdm(self.获取课次列表()):
            resource_id = lt[i]
            formatted_date = self.today  # datetime.datetime.now().strftime("%Y-%m-%d")
            path = prfx.format(x=i, y=formatted_date)
            self.获取直播间用户数据(resource_id, self.root / "数据表" / path)

    # 20240206 新增【针对打卡部分
    def 获取打卡id(self):
        """ 依据打卡链接，获取activity_id(打卡id)
        """
        打卡链接 = self.打卡链接[1:]
        ls_activity_id = []
        for item in 打卡链接:  # 课次
            activity_id = re.search(r"\?activity_id=(.+?)\&", item)
            ls_activity_id.append(activity_id.group(1))
        return ls_activity_id

    def 获取打卡参与用户(self, activity_id, path):
        # 1）如果路径中已经有了，就跳过
        if path.is_file():
            return
            # 2）获取打卡用户数据：
        lst = self.xe.get_elock_actor(activity_id)
        fieldnames = ['用户id', '用户昵称', '打卡昵称', '打卡分组', '姓名', '电话', '最近采集号码', '城市', '微信号',
                      '打卡天数', '打卡次数', '被点赞数', '被评论数', '被点评数', '被精选数', '参与时间']
        p = path
        with open(p, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            writer.writeheader()
            for x in lst:
                record = {
                    '用户id': x['user_id'],
                    '用户昵称': x['wx_nickname'],
                    '打卡昵称': x['clock_nickname'],
                    '打卡分组': None,
                    '姓名': x['wx_nickname'],
                    '电话': x['phone'],
                    '最近采集号码': None,
                    '城市': x['wx_city'],
                    '微信号': None,
                    '打卡天数': x['clock_days'],
                    '打卡次数': x['clock_days'],
                    '被点赞数': x['zan_count'],
                    '被评论数': x['comment_count'],
                    '被点评数': x['review_count'],
                    '被精选数': 0,
                    '参与时间': x['created_at']
                }
                writer.writerow(record)

    def 下载打卡数据(self):
        prfx = "{x}-" + f"《{self.返款标题}技术公益网课【中心教室】-日历打卡学员数据.csv"
        lt = self.获取打卡id()
        for i in tqdm(range(1)):
            activity_id = lt[i]
            formatted_date = self.today  # datetime.datetime.now().strftime("%Y-%m-%d")
            path = prfx.format(x=formatted_date)
            self.获取打卡参与用户(activity_id, self.root / "数据表" / path)


class 课次数据:
    """ 读取课次数据 """

    def __init__(self):
        # 将各种不同来源的数据，按用户ID分组存储起来
        self.用户观看数据 = defaultdict(list)  # 每个用户的观看数据情况
        self.start_day = None  # 记录所有加载文件中，最早的时间戳，大概率就是这个课次的开课时间

    def add_files(self, path, pattern):
        files = XlPath(path).select_file(pattern)
        for f in files:
            self.add_考勤数据(f)

    def add_考勤数据(self, file):
        """ 推荐使用这个综合接口，代替专用接口，这个接口能实现一些综合性的操作
        保存的用户数据，存储三元数值（观看时长，进度）
            150，表示观看时间满30分钟
            0~100，表示有进度数据，0%~100%
        """
        # 0 有普通考勤表，和从圈子下载的表格两种
        enc = get_encoding(file.read_bytes()) or 'gbk'
        df = pd.read_csv(file, encoding=enc, encoding_errors='ignore')

        if '参与状态' in df.columns:
            self.add_圈子进度表(file)
        else:
            self.add_小鹅通考勤表(file)

        dates = re.findall(r'(\d{4}-\d{2}-\d{2})_\d{2}-\d{2}-\d{2}', file.stem)
        if dates:
            file_day = datetime.datetime.strptime(dates[-1], "%Y-%m-%d").date()
            if self.start_day is None or self.start_day > file_day:
                self.start_day = file_day

    def add_小鹅通考勤表(self, file):
        file = XlPath(file)

        df = pd.read_csv(file, encoding='utf8')
        assert '直播观看时长(秒)' in df.columns, f'好像下载到空数据表 {file}'

        for idx, x in df.iterrows():
            data = {'文件名': XlPath(file).stem}
            data['直播分钟'] = int(x['直播观看时长(秒)']) // 60
            data['回放分钟'] = int(x['回放观看时长(秒)']) // 60
            self.用户观看数据[x['用户ID']].append(data)

    def add_圈子进度表(self, file):
        """ 圈子的考勤数据比较特别，需要独立的一个体系来存储
        大部分会有播放进度，
        """
        file = XlPath(file)
        # 圈子数据目前是gbk编码，这样会有些问题。但可能哪天平台就改成utf8编码了，所以这里提前做好兼容。
        enc = get_encoding(file.read_bytes()) or 'gbk'
        df = pd.read_csv(file, encoding=enc, encoding_errors='ignore')

        for idx, x in df.iterrows():
            data = {'文件名': file.stem}

            if '播放进度' in x:  # 一般会有详细的播放进度
                data['百分比进度'] = int(re.search(r'\d+', x['播放进度']).group())
            else:  # 如果没有，也可以通过参与状态知道是否完成
                data['百分比进度'] = 100 if x['参与状态'] == '已完成' else 0
            # 虽然"已完成"，但显示进度可能不是'100%'，要补一个修正
            if x['参与状态'] == '已完成':
                data['百分比进度'] = 100

            self.用户观看数据[x['用户ID']].append(data)

    def 禅宗考勤结果(self, user_id):
        """ 禅宗考勤是比较简单的，不用考虑回放情况，只要在所有数据中，判断出有完成就行
        当然，细节考虑，也还是可以详细区分不同情况。

        :return:
            str，一般是返回字符串，表示结论
            dict，但也可以返回一个字典，表示特殊的格式设置，这种是用于单元格颜色设置
        """
        直播分钟, 回放分钟, 百分比进度 = 0, 0, 0

        if user_id not in self.用户观看数据:
            return {'value': f'未开始', 'color': '白色'}

        # 1 遍历所有文件数据，获得所有最大值
        data = self.用户观看数据[user_id]  # 这个类的功能框架，已经按照user_id进行了数据分组统计
        for x in data:
            # 前两条是兼容以前的通用的直播课程数据
            直播分钟 = max(直播分钟, x.get('直播分钟', 0))
            回放分钟 = max(回放分钟, x.get('回放分钟', 0))
            # 这一条是兼容禅宗特有的百分比进度
            百分比进度 = max(百分比进度, x.get('百分比进度', 0))

        # 2 归纳出本课次考勤结论
        yellow = RgbFormatter.from_name('鲜黄')
        if 百分比进度 == 100 or 回放分钟 >= 30:
            return {'value': '已完成', 'color': '鲜绿色'}
        elif 百分比进度:
            return {'value': f'进度{百分比进度}%', 'color': yellow.light((100 - 百分比进度) / 100)}
        elif 回放分钟:
            return {'value': f'观看{回放分钟}分钟', 'color': yellow.light((100 - 回放分钟) / 30)}
        else:
            return {'value': f'未开始', 'color': '白色'}

    def 小鹅通考勤结果(self, user_id, 需求分钟=30):
        """ 一般是把一个课次的多天数据添加到一起用，用这个接口获得该课次的观看状态
        比如是当堂，还是第1天回放等

        :param user_id: 要获得的用户信息
        :param 需求分钟: 判断完成的依据分钟数
        """
        # 1 检查文件日期
        if user_id not in self.用户观看数据:
            return '未开始学习'

        data = self.用户观看数据[user_id]

        # 理论上这里文件名就应该是排好序的
        filenames = [x['文件名'] for x in data]

        for i, filename in enumerate(filenames):
            # 后面一串是时分秒，并不提取，但是用来限定格式匹配，避免匹配错
            dates = re.findall(r'(\d{4}-\d{2}-\d{2})_\d{2}-\d{2}-\d{2}', filename)
            if not dates:  # 找不到时间戳的不处理
                continue

            file_date_obj = datetime.datetime.strptime(dates[0], "%Y-%m-%d").date()
            delta_day = (file_date_obj - self.start_day).days

            # 2 按顺序提取时间
            if delta_day == 0 and data[i]['直播分钟'] >= 需求分钟:
                return '完成当堂学习'
            elif delta_day > 0 and data[i]['直播分钟'] + data[i]['回放分钟'] >= 需求分钟:
                return f'第{delta_day}天回放'

        if data[-1]['直播分钟'] + data[-1]['回放分钟']:
            return f'不足{需求分钟}分钟'
        else:
            return '未开始学习'

    def 小鹅通考勤结果2(self, user_id, 返款梯度, 要求在线分钟=30):
        """ 输入视频返款的梯度，这个函数相比`小鹅通考勤结果`，还会返回单元格颜色格式，对应的返款额 """
        # 1 判断课程回放是不是结束了
        text = self.小鹅通考勤结果(user_id, 要求在线分钟)
        if text == '未开始学习' or '不足' in text:  # 判断该课次是否已经结束了
            delta_day = datetime.timedelta(days=len(返款梯度))
            if (delta_day + self.start_day) <= datetime.date.today():
                text = '未完成学习'

        if text == '完成当堂学习':
            return text, '鲜绿色', 返款梯度[0]
        elif m := re.match(r'第(\d+)天回放', text):
            t = int(m.group(1))
            money = 返款梯度[t]
            if money:
                color = RgbFormatter.from_name('黄色')
                color = color.light((返款梯度[0] - money) / money)  # 根据返款额度自动变浅
            else:
                color = '灰色'
            return text, color, money
        elif text == '未开始学习' or '不足' in text:
            return text, '白色', 0
        elif text == '未完成学习':
            return text, '红色', 0
        else:
            raise ValueError


def 聚合读取考勤数据(data_dir, name, judge_minute=30):
    """
    1、在目录"data_folder"下，找前缀包含name的所有文件
    2、并将其按日期排序，整理考勤数据，判断当堂、第一天完成等
    3、课次完成的标记，是judge_minute要满足特定分钟数
    """
    # 理论上获得的第一份就是当堂数据，第二份则是回放数据
    files = list(XlPath(data_dir).rglob_files(f'*{name}*'))
    data = 课次数据()
    for f in files:
        data.add_考勤数据(f)

    for user_id in data.用户观看数据:
        print(data.小鹅通考勤结果2(user_id, [100, 90, 80, 70, 0, 0]))
