#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Author : 陈坤泽
# @Email  : 877362867@qq.com
# @Data   : 2020/06/02


"""
扩展了些自己的openpyxl工具
"""


# 可能缺openpyxl，需要执行这个先安装
from pyxllib.util.tablepyxl import tablepyxl

import openpyxl
from openpyxl.styles import Font


from pyxllib.debug.all import *


class Openpyxl:
    """
    对openpyxl库做一些基本的功能拓展
    """

    @staticmethod
    def address(n, m) -> str:
        r"""数字索引转excel地址索引
        :param n: 行号，可以输入字符串形式的数字
        :param m: 列号，同上可以输入str的数字
        :return:

        >>> Openpyxl.address(2, 3)
        'C2'
        """
        from openpyxl.utils.cell import get_column_letter
        return f'{get_column_letter(int(m))}{n}'

    @staticmethod
    def in_range(cell):
        """判断一个单元格所在的合并单元格
        >> in_range(ws['C1'])
        <openpyxl.worksheet.cell_range.CellRange> A1:D3
        """
        ws = cell.parent
        for rng in ws.merged_cells.ranges:
            if cell.coordinate in rng:
                break
        else:  # 如果找不到则返回原值
            rng = cell
        return rng

    @staticmethod
    def mcell(cell):
        """返回“有效单元格”，即如果输入的是一个合并单元格，会返回该合并单元格左上角的单元格
        修改左上角单元格的值才是可行、有意义的

        因为跟合并单元格有关，所以 以m前缀 merge
        """
        from openpyxl.cell.cell import MergedCell
        if isinstance(cell, MergedCell):
            ws = cell.parent
            xy = Openpyxl.in_range(cell).top[0]
            return ws[Openpyxl.address(*xy)]
        else:
            return cell

    @staticmethod
    def celltype(cell):
        """
        :param cell: 一个单元格
        :return: 单元格类型
            0：普通单元格
            1：合并单元格其他衍生位置
            2：合并单元格的左上角的位置

        TODO 这个函数还是可以看看能不能有更好的实现、提速
        """
        from openpyxl.cell.cell import MergedCell
        if isinstance(cell, MergedCell):
            return 1
        elif isinstance(cell.offset(1, 0), MergedCell) or isinstance(cell.offset(0, 1), MergedCell):
            # 这里只能判断可能是合并单元格，具体是不是合并单元格，还要
            rng = Openpyxl.in_range(cell)
            return 2 if hasattr(rng, 'size') else 0
        else:
            return 0

    @staticmethod
    def isnone(cell):
        """是普通单元格且值为None
        注意合并单元格的衍生单元格不为None
        """
        celltype = Openpyxl.celltype(cell)
        return celltype == 0 and cell.value is None

    @staticmethod
    def copy_cell_format(cell, new_cell):
        """ 单元格全格式复制，需要事先指定好新旧单元格的物理位置
        参考：https://stackoverflow.com/questions/23332259/copy-cell-style-openpyxl
        """
        from copy import copy
        if cell.has_style:
            new_cell.font = copy(cell.font)  # 字体
            new_cell.border = copy(cell.border)  # 表格线
            new_cell.fill = copy(cell.fill)  # 填充色
            new_cell.number_format = copy(cell.number_format)  # 数字格式
            new_cell.protection = copy(cell.protection)  # 保护？
            new_cell.alignment = copy(cell.alignment)  # 对齐格式
            # new_cell.style = cell.style
        # if cell.comment:
            # 这个会引发AttributeError。。。
            #       vml = fromstring(self.workbook.vba_archive.read(ws.legacy_drawing))
            #   AttributeError: 'NoneType' object has no attribute 'read'
            # new_cell.comment = copy(cell.comment)
            # 就算开了keep_vba可以强制写入了，打开的时候文件可能还是会错

    @staticmethod
    def copy_cell(cell, new_cell):
        """ 单元格全格式、包括值的整体复制
        """
        new_cell.value = cell.value
        Openpyxl.copy_cell_format(cell, new_cell)

    @classmethod
    def down(cls, cell):
        """输入一个单元格，向下移动一格
        注意其跟offset的区别，如果cell是合并单元格，会跳过自身的衍生单元格
        """
        if cls.celltype(cell):  # 合并单元格
            rng = cls.in_range(cell)
            return cell.parent.cell(rng.max_row+1, cell.column)
        else:
            return cell.offset(1, 0)

    @classmethod
    def right(cls, cell):
        if cls.celltype(cell):
            rng = cls.in_range(cell)
            return cell.parent.cell(cell.row, rng.max_row+1)
        else:
            return cell.offset(0, 1)

    @classmethod
    def up(cls, cell):
        if cls.celltype(cell):
            rng = cls.in_range(cell)
            return cell.parent.cell(rng.min_row-1, cell.column)
        else:
            return cell.offset(-1, 0)

    @classmethod
    def left(cls, cell):
        if cls.celltype(cell):
            rng = cls.in_range(cell)
            return cell.parent.cell(cell.row, rng.min_row-1)
        else:
            return cell.offset(0, -1)

    @staticmethod
    def copy_worksheet(origin_ws, target_ws):
        """跨工作薄时复制表格内容的功能
        openpyxl自带的Workbook.copy_worksheet没法跨工作薄复制，很坑
        """
        # 1、取每个单元格的值
        for row in origin_ws:
            for cell in row:
                try:
                    Openpyxl.copy_cell(cell, target_ws[cell.coordinate])
                except AttributeError:
                    pass
        # 2、合并单元格的处理
        for rng in origin_ws.merged_cells.ranges:
            target_ws.merge_cells(rng.ref)
        # 3、其他表格属性的复制
        # 这个从excel读取过来的时候，是不准的，例如D3可能因为关闭时停留窗口的原因误跑到D103
        # dprint(origin_ws.freeze_panes)
        # target_ws.freeze_panes = origin_ws.freeze_panes


def product(*iterables, order=None, repeat=1):
    """ 对 itertools 的product扩展orders参数的更高级的product迭代器
    :param order: 假设iterables有n=3个迭代器，则默认 orders=[1, 2, 3] （起始编号1）
        即标准的product，是按顺序对每个迭代器进行重置、遍历的
        但是我扩展的这个接口，允许调整每个维度的更新顺序
        例如设置为 [-2, 1, 3]，表示先对第2维降序，然后按第1、3维的方式排序获得各个坐标点
        注：可以只输入[-2]，默认会自动补充维[1, 3]

    for x in product('ab', 'cd', 'ef', order=[3, -2, 1]):
        print(x)

    ['a', 'd', 'e']
    ['b', 'd', 'e']
    ['a', 'c', 'e']
    ['b', 'c', 'e']
    ['a', 'd', 'f']
    ['b', 'd', 'f']
    ['a', 'c', 'f']
    ['b', 'c', 'f']

    TODO 我在想numpy这么牛逼，会不会有等价的功能接口可以实现，我不用重复造轮子？
    """
    import itertools, numpy

    # 一、标准调用方式
    if order is None:
        for x in itertools.product(*iterables, repeat=repeat):
            yield x
        return

    # 二、输入orders参数的调用方式
    # 1、补全orders参数长度
    n = len(iterables)
    for i in range(1, n + 1):
        if not (i in order or -i in order):
            order.append(i)
    if len(order) != n: return ValueError(f'orders参数值有问题 {order}')

    # 2、生成新的迭代器组
    new_iterables = [(iterables[i-1] if i > 0 else reversed(iterables[-i-1])) for i in order]
    idx = numpy.argsort([abs(i) - 1 for i in order])
    for y in itertools.product(*new_iterables, repeat=repeat):
        yield [y[i] for i in idx]


class Worksheet(openpyxl.worksheet.worksheet.Worksheet):
    """ 扩展标准的Workshhet功能
    >> wb = openpyxl.load_workbook(filename='高中数学知识树匹配终稿.xlsx', data_only=True)
    >> ws1 = Worksheet(wb['main'])
    >> ws2 = Worksheet(wb['导出'])
    """

    def __init__(self, ws):
        self.__dict__ = ws.__dict__

    def _cells_by_row(self, min_col, min_row, max_col, max_row, values_only=False):
        """openpyxl的这个迭代器，遇到合并单元格会有bug
        所以我把它重新设计一下~~
        """
        for row in range(min_row, max_row + 1):
            cells = (self.cell(row=row, column=column) for column in range(min_col, max_col + 1))
            if values_only:
                # yield tuple(cell.value for cell in cells)  # 原代码
                yield tuple(getattr(cell, 'value', None) for cell in cells)
            else:
                yield tuple(cells)

    def search(self, pattern, min_row=None, max_row=None, min_col=None, max_col=None, order=None, direction=0):
        """查找满足pattern正则表达式的单元格

        :param pattern: 正则匹配式，可以输入re.complier对象
            会将每个单元格的值转成str，然后进行字符串规则search匹配
            支持多层嵌套 ['模块一', '属性1']
        :param direction: 只有在 pattern 为数组的时候有用
            pattern有多组时，会嵌套找单元格
            每计算出一个条件后，默认取该单元格下方的子区间 axis=0
            如果不是下方，而是右方，可以设置为1
            还有奇葩的上方、左方，可以分别设置为2、3
        :param order: 默认None，也就是 [1, 2] 的效果，规律详见product接口

        >> wb = openpyxl.load_workbook(filename='2020寒假教材各地区数量统计最新2020.1.1.xlsx')
        >> ws = Worksheet(wb['预算总表'])
        >> ws.search('年段')
        <Cell '预算总表'.B2>
        """
        # 1、定界
        x1, x2 = max(min_row or 1, 1), min(max_row or self.max_row, self.max_row)
        y1, y2 = max(min_col or 1, 1), min(max_col or self.max_column, self.max_column)

        # 2、遍历
        if isinstance(pattern, (list, tuple)):
            cel = None
            for p in pattern:
                cel = self.search(p, x1, x2, y1, y2, order)
                if cel:
                    # up, down, left, right 找到的单元格四边界
                    l, u, r, d = getattr(Openpyxl.in_range(cel), 'bounds', (cel.column, cel.row, cel.column, cel.row))
                    if direction == 0:
                        x1, y1, y2 = max(x1, d), max(y1, l), min(y2, r)
                    elif direction == 1:
                        x1, x2, y1 = max(x1, u), min(x2, d), max(y1, r)
                    elif direction == 2:
                        x2, y1, y2 = min(x2, d), max(y1, l), min(y2, r)
                    elif direction == 3:
                        x1, x2, y2 = max(x1, u), min(x2, d), min(y2, l)
                    else:
                        raise ValueError(f'direction参数值错误{direction}')
                else:
                    return None
            return cel
        else:
            if isinstance(pattern, str): pattern = re.compile(pattern)
            for x, y in product(range(x1, x2 + 1), range(y1, y2 + 1), order=order):
                cell = self.cell(x, y)
                if Openpyxl.celltype(cell) == 1: continue  # 过滤掉合并单元格位置
                if pattern.search(str(cell.value)): return cell  # 返回满足条件的第一个值

    findcel = search

    def findrow(self, pattern, *args, **kwargs):
        cel = self.findcel(pattern, *args, **kwargs)
        return cel.row if cel else 0

    def findcol(self, pattern, *args, **kwargs):
        cel = self.findcel(pattern, *args, **kwargs)
        return cel.column if cel else 0

    def chrome(self):
        """注意，这里会去除掉合并单元格"""
        chrome(pd.DataFrame(self.values))

    def select_columns(self, columns, column_name='searchkey'):
        r"""获取表中columns属性列的值，返回dataframe数据类型

        :param columns: 搜索列名使用正则re.search字符串匹配查找
            可以单列：'attr1'，找到列头后，会一直往后取到最后一个非空值
            也可以多列： ['attr1', 'attr2', 'attr3']
                会结合多个列标题定位，数据从最大的起始行号开始取，
                （TODO 截止到最末非空值所在行  未实现，先用openpyxl自带的max_row判断，不过这个有时会判断过大）
            遇到合并单元格，会寻找其母单元格的值填充
        :param column_name: 返回的df。列名
            origin，原始的列名
            searchkey，搜索时用的查找名
        """
        if not isinstance(columns, (list, tuple)):
            columns = [columns]

        # 1、找到所有标题位置，定位起始行
        cels, names, start_line = [], [], -1
        for search_name in columns:
            cel = self.findcel(search_name)
            if cel:
                cels.append(cel)
                if column_name=='searchkey':
                    names.append(str(search_name))
                elif column_name=='origin':
                    if isinstance(search_name, (list, tuple)) and len(search_name) > 1:
                        names.append('/'.join(list(search_name[:-1]) + [str(cel.value)]))
                    else:
                        names.append(str(cel.value))
                else:
                    raise ValueError(f'{column_name}')
                start_line = max(start_line, Openpyxl.down(cel).row)
            else:
                dprint(search_name)  # 找不到指定列

        # 2、获得每列的数据
        datas = {}
        for k, cel in enumerate(cels):
            if cel:
                col = cel.column
                li = []
                for i in range(start_line, self.max_row+1):
                    v = Openpyxl.mcell(self.cell(i, col)).value  # 注意合并单元格的取值
                    li.append(v)
                datas[names[k]] = li
            else:
                # 如果没找到列，设一个空列
                datas[names[k]] = [None]*(self.max_row + 1 - start_line)
        df = pd.DataFrame(datas)

        # 3、去除所有空行数据
        df.dropna(how='all', inplace=True)

        return df

    def copy_range(self, cell_range, rows=0, cols=0):
        """ 同表格内的 range 复制操作
        Copy a cell range by the number of rows and/or columns:
        down if rows > 0 and up if rows < 0
        right if cols > 0 and left if cols < 0
        Existing cells will be overwritten.
        Formulae and references will not be updated.
        """
        from openpyxl.worksheet.cell_range import CellRange
        from itertools import product
        # 1、预处理
        if isinstance(cell_range, str):
            cell_range = CellRange(cell_range)
        if not isinstance(cell_range, CellRange):
            raise ValueError("Only CellRange objects can be copied")
        if not rows and not cols:
            return
        min_col, min_row, max_col, max_row = cell_range.bounds
        # 2、注意拷贝顺序跟移动方向是有关系的，要防止被误覆盖，复制了新的值，而非原始值
        r = sorted(range(min_row, max_row+1), reverse=rows > 0)
        c = sorted(range(min_col, max_col+1), reverse=cols > 0)
        for row, column in product(r, c):
            Openpyxl.copy_cell(self.cell(row, column), self.cell(row + rows, column+cols))

    def reindex_columns(self, orders):
        """ 重新排列表格的列顺序
        >> ws.reindex_columns('I,J,A,,,G,B,C,D,F,E,H,,,K'.split(','))

        TODO 支持含合并单元格的整体移动？
        """
        from openpyxl.utils.cell import column_index_from_string
        max_row, max_column = self.max_row, self.max_column
        for j, col in enumerate(orders, 1):
            if not col: continue
            self.copy_range(f'{col}1:{col}{max_row}', cols=max_column+j-column_index_from_string(col))
        self.delete_cols(1, max_column)


def adjust_sheets(wb, new_sheetnames):
    """ 按照 new_sheetnames 的清单重新调整sheets
        在清单里的按顺序罗列
        不在清单里的表格删除
        不能出现wb原本没有的表格名
    """
    for name in set(wb.sheetnames) - set(new_sheetnames):
        # 最好调用标准的remove接口删除sheet
        #   不然虽然能表面上看也能删除sheet，但会有命名空间的一些冗余信息留下
        wb.remove(wb[name])
    wb._sheets = [wb[name] for name in new_sheetnames]
    return wb


def demo_openpyxl():
    # 一、新建一个工作薄
    from openpyxl import Workbook
    wb = Workbook()

    # 取一个工作表
    ws = wb.active  # wb['Sheet']，取已知名称、下标的表格，excel不区分大小写，这里索引区分大小写

    # 1、索引单元格的两种方法，及可以用.value获取值
    ws['A2'] = '123'
    dprint(ws.cell(2, 1).value)  # 123

    # 2、合并单元格
    ws.merge_cells('A1:C2')
    dprint(ws['A1'].value)  # None，会把原来A2的内容清除

    # print(ws['A2'].value)  # AttributeError: 'MergedCell' object has no attribute 'value'

    # ws.unmerge_cells('A1:A3')  # ValueError: list.remove(x): x not in list，必须标记完整的合并单元格区域，否则会报错
    ws['A1'].value = '模块一'
    ws['A3'].value = '属性1'
    ws['B3'].value = '属性2'
    ws['C3'].value = '属性3'

    ws.merge_cells('D1:E2')
    ws['D1'].value = '模块二'
    ws['D3'].value = '属性1'
    ws['E3'].value = '属性2'

    dprint(ws['A1'].offset(1, 0).coordinate)  # A2
    dprint(Openpyxl.down(ws['A1']).coordinate)  # A3

    # 3、设置单元格样式、格式
    from openpyxl.comments import Comment
    cell = ws['A3']
    cell.font = Font(name='Courier', size=36)
    cell.comment = Comment(text="A comment", author="Author's Name")
    from openpyxl.styles.colors import RED

    styles = [['Number formats', 'Comma', 'Comma [0]', 'Currency', 'Currency [0]', 'Percent'],
              ['Informative', 'Calculation', 'Total', 'Note', 'Warning Text', 'Explanatory Text'],
              ['Text styles', 'Title', 'Headline 1', 'Headline 2', 'Headline 3', 'Headline 4', 'Hyperlink', 'Followed Hyperlink', 'Linked Cell'],
              ['Comparisons', 'Input', 'Output', 'Check Cell', 'Good', 'Bad', 'Neutral'],
              ['Highlights', 'Accent1', '20 % - Accent1', '40 % - Accent1', '60 % - Accent1', 'Accent2', 'Accent3', 'Accent4', 'Accent5', 'Accent6', 'Pandas']]
    for i, name in enumerate(styles, start=4):
        ws.cell(i, 1, name[0])
        for j, v in enumerate(name[1:], start=2):
            ws.cell(i, j, v)
            ws.cell(i, j).style = v

    # 二、测试一些功能
    ws = Worksheet(ws)

    dprint(ws.search('模块二').coordinate)  # D1
    dprint(ws.search(['模块二', '属性1']).coordinate)  # D3

    dprint(ws.findcol(['模块一', '属性1'], direction=1))  # 0

    wb.save("demo_openpyxl.xlsx")
