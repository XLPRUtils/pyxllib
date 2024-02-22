#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Author : 陈坤泽
# @Email  : 877362867@qq.com
# @Date   : 2024/01/07

"""
扩展了些自己的openpyxl工具
"""
import time

from pyxllib.prog.pupil import check_install_package, run_once

check_install_package('openpyxl')
check_install_package('premailer')
# check_install_package('xlrd2')
check_install_package('yattag')
check_install_package('jsonpickle')

from collections import Counter, OrderedDict, defaultdict
import csv
import datetime
from itertools import islice
import json
import math
from pathlib import Path
import random
import re
import io

import xlrd

import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment
from openpyxl.utils.cell import get_column_letter, column_index_from_string
import pandas as pd

try:
    import jsonpickle
except ModuleNotFoundError:
    pass

from pyxllib.prog.newbie import human_readable_number
from pyxllib.prog.pupil import (inject_members, dprint, xlmd5, shuffle_dict_keys, Timeout,
                                safe_div, format_exception, DictTool)
from pyxllib.prog.specialist import browser, TicToc
from pyxllib.algo.specialist import product
from pyxllib.text.pupil import calc_chinese_ratio
from pyxllib.file.specialist import XlPath


def __1_basic():
    """ 表格的组件功能 """


def excel_addr(n, m) -> str:
    r"""数字索引转excel地址索引

    :param n: 行号，可以输入字符串形式的数字
    :param m: 列号，同上可以输入str的数字
    :return:

    >>> excel_addr(2, 3)
    'C2'
    """
    return f'{get_column_letter(int(m))}{n}'


def excel_addr2(n1, m1, n2, m2) -> str:
    r""" excel_addr的扩展版，定位一个区间

    >>> excel_addr2(2, 3, 4, 4)
    'C2:D4'
    """
    return f'{get_column_letter(int(m1))}{n1}:{get_column_letter(int(m2))}{n2}'


def is_valid_excel_cell(cell):
    """ 判断输入的字符串是否是一个合法的Excel单元格地址

    :param str cell: 输入的字符串
    :return bool: 如果是合法的Excel单元格地址返回True，否则返回False
    """
    match = re.fullmatch(r'[A-Z]+[1-9][0-9]*', cell)
    return match is not None


def is_valid_excel_range(range):
    """ 判断输入的字符串是否是一个合法的Excel单元格范围

    :param str range: 输入的字符串
    :return bool: 如果是合法的Excel单元格范围返回True，否则返回False
    """
    if ':' in range:
        start, end = range.split(':')
        return (is_valid_excel_cell(start) or start.isdigit() or re.fullmatch(r'[A-Z]+', start)) and \
            (is_valid_excel_cell(end) or end.isdigit() or re.fullmatch(r'[A-Z]+', end)) and \
            start <= end
    else:
        return is_valid_excel_cell(range)


def is_valid_excel_address(address):
    """ 判断输入的字符串是否是一个合法的Excel地址定位

    :param str address: 输入的字符串
    :return bool: 如果是合法的Excel地址定位返回True，否则返回False

    注意，严格来说，'A1,A3'这种定位也是可以的，但是这个函数暂不考虑这种情况，
        如果需要，可以另外写is_valid_excel_address2
    """
    if ':' in address:
        return is_valid_excel_range(address)
    else:
        return is_valid_excel_cell(address)


@run_once('str', debug=True)
def xlfmt2pyfmt_date(xl_fmt):
    """ 日期的渲染操作

    >>> xlfmt2pyfmt_date('yyyy/m/d')
    '%Y/{month}/{day}'
    >>> xlfmt2pyfmt_date('yyyy-mm-dd')
    '%Y-%m-%d'
    >>> xlfmt2pyfmt_date('yyyy年mm月dd日')
    '%Y年%m月%d日'
    >>> xlfmt2pyfmt_date('yyyy年m月d日')
    '%Y年{month}月{day}日'

    # 注意以下是并未支持的功能...会默认返回 '%Y/%-m/%-d'
    >>> xlfmt2pyfmt_date('yy-m-d')
    '%y-{month}-{day}'
    >>> xlfmt2pyfmt_date('m/d/yy')
    '{month}/{day}/%y'
    >>> xlfmt2pyfmt_date('dddd, mmmm dd, yyyy')
    '%A, %B {day}, %Y'
    >>> xlfmt2pyfmt_date('yy/mm/dd')
    '%y/%m/%d'
    >>> xlfmt2pyfmt_date('m-d-yy')
    '{month}-{day}-%y'
    """
    mappings = {
        'y': {2: '%y', 4: '%Y'},
        'm': {1: '%-m', 2: '%m', 3: '%b', 4: '%B'},
        'd': {1: '%-d', 2: '%d', 3: '%a', 4: '%A'}
    }

    m = re.search(r'(y+)(.+?)(m+)(.+?)(d+)(日?)', xl_fmt.replace('"', ''))
    if m:
        y, sep1, m, sep2, d, sep3 = m.groups()
        year_pattern = mappings['y'].get(len(y), '%Y')
        month_pattern = mappings['m'].get(len(m), '%m')
        day_pattern = mappings['d'].get(len(d), '%d')
        fmt = f'{year_pattern}{sep1}{month_pattern}{sep2}{day_pattern}{sep3}'
    else:
        fmt = '%Y/%-m/%-d'

    # 在windows下，%-m和%-d会报错，所以需要替换成{month}和{day}
    fmt = fmt.replace('%-m', '{month}')
    fmt = fmt.replace('%-d', '{day}')
    return fmt


@run_once('str')
def xlfmt2pyfmt_time(xl_fmt):
    """ 时间的渲染操作

    >>> xlfmt2pyfmt_time('h:mm:ss')
    '%I:%M:%S'
    >>> xlfmt2pyfmt_time('hh:mm:ss')
    '%H:%M:%S'
    >>> xlfmt2pyfmt_time('mm:ss')
    '%M:%S'
    >>> xlfmt2pyfmt_time('h:mm')
    '%I:%M'
    >>> xlfmt2pyfmt_time('hh:mm')
    '%H:%M'
    >>> xlfmt2pyfmt_time('m:ss')
    '%M:%S'
    >>> xlfmt2pyfmt_time('h:mm:ss AM/PM')
    '%I:%M:%S %p'
    >>> xlfmt2pyfmt_time('hh:mm:ss AM/PM')
    '%H:%M:%S %p'
    """
    xl_fmt = re.sub(r'(y+)(.+?)(m+)(.+?)(d+)(日?)', '', xl_fmt.replace('"', ''))

    components = []

    # 判断是12小时制还是24小时制
    if 'hh' in xl_fmt:
        components.append('%H')
    elif 'h' in xl_fmt:
        components.append('%I')

    # 判断是否显示分钟
    if 'mm' in xl_fmt or 'm' in xl_fmt:
        components.append('%M')

    # 判断是否显示秒钟
    if 'ss' in xl_fmt or 's' in xl_fmt:
        components.append('%S')

    # 判断是否显示AM/PM
    if 'AM/PM' in xl_fmt:
        if components:
            components[-1] += ' %p'
        else:
            components.append('%p')

    return ':'.join(components)


# @run_once('str')
def xlfmt2pyfmt_datetime(xl_fmt):
    """ 主要是针对日期、时间的渲染操作

    >>> xlfmt2pyfmt_datetime('yyyy-mm-dd h:mm:ss')
    '%Y-%m-%d %I:%M:%S'
    """
    py_fmt = xlfmt2pyfmt_date(xl_fmt)
    if ':' in xl_fmt:
        py_fmt += ' ' + xlfmt2pyfmt_time(xl_fmt)
    return py_fmt


def xl_render_value(x, xl_fmt):
    """ 得到单元格简单渲染后的效果
    py里不可能对excel的所有格式进行全覆盖，只是对场景格式进行处理

    注意，遇到公式是很难计算处理的，大概率只能保持原公式显示
    因为日期用的比较多，需要时常获得真实的渲染效果，所以这里封装一个接口

    >>> xl_render_value(datetime.datetime(2020, 1, 1), 'yyyy-mm-dd')
    '2020-01-01'
    """

    if isinstance(x, datetime.datetime):
        y = x.strftime(xlfmt2pyfmt_datetime(xl_fmt)).format(month=x.month, day=x.day)
    elif isinstance(x, datetime.date):
        y = x.strftime(xlfmt2pyfmt_date(xl_fmt)).format(month=x.month, day=x.day)
    elif isinstance(x, datetime.time):
        y = x.strftime(xlfmt2pyfmt_time(xl_fmt))
    elif isinstance(x, datetime.timedelta):
        y = str(x)
    elif isinstance(x, (str, int, float, bool)):  # 其他可以json化的数据类型
        y = x
    else:  # ArrayFormula、DataTableFormula等无法json化的数据，提前转str
        y = str(x)
    return y


def sort_excel_files(file_paths):
    """ 在文件清单中，把excel类型的文件优先排到前面 """

    def sort_key(filename: str) -> int:
        """ 根据文件后缀给出权重排序值

        :param str filename: 文件名
        :return int: 权重值（小的在前）

        >>> sort_key('test.xlsx')
        1
        >>> sort_key('demo.xls')
        2
        >>> sort_key('other.txt')
        3
        """
        if re.search(r'\.xlsx$', filename):
            return 1, filename
        elif re.search(r'\.xl[^.]*$', filename):
            return 2, filename
        else:
            return 3, filename

    file_paths2 = sorted(file_paths, key=sort_key)
    return file_paths2


def excel2md5(file, reduction_degree=1):
    try:
        wb = openpyxl.load_workbook(file)
    except (ValueError, TypeError) as e:
        # 有些表格直接读取会失败，但使用read_only就能读了
        wb = openpyxl.load_workbook(file, read_only=True)
    except Exception as e:  # 还有其他zipfile.BadZipFile等错误
        print(file, str(e))
        return ''

    return wb.to_md5(reduction_degree)


def convert_csv_text_to_xlsx(csv_text):
    """ 将 csv 文本转换为 xlsx 文件 """
    wb = Workbook()
    sheet = wb.active

    # 使用 io.StringIO 将 csv 文本转换为可读的文件对象
    f = io.StringIO(csv_text)
    reader = csv.reader(f)
    for row_idx, row in enumerate(reader, start=1):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx).value = value

    f.close()
    return wb


def convert_csv_to_xlsx(csv_file):
    """ 将 csv 文件转换为 xlsx 文件 """
    wb = Workbook()
    sheet = wb.active

    with open(csv_file, encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx).value = value

    return wb


def convert_xls_to_xlsx(xls_file):
    """ 将 xls 文件转换为 xlsx 文件

    注意，这只是一个简化版的转换，要尽量完整的话，还是要用microsoft 365来升级xls的
    """
    # 使用 xlrd 打开 xls 文件
    xls_workbook = xlrd.open_workbook(xls_file)

    # 创建一个新的 openpyxl 工作簿
    wb = Workbook()
    ws = wb.active

    for i in range(xls_workbook.nsheets):
        xls_sheet = xls_workbook.sheet_by_index(i)
        if i == 0:
            # 为第一个工作表设置名称
            ws.title = xls_sheet.name
        else:
            # 创建新的工作表并设置名称
            ws = wb.create_sheet(title=xls_sheet.name)
        for row in range(xls_sheet.nrows):
            for col in range(xls_sheet.ncols):
                # 将 xlrd 单元格的数据写入 openpyxl 单元格
                ws.cell(row=row + 1, column=col + 1).value = xls_sheet.cell_value(row, col)

    return wb


def load_as_xlsx_file(file_path, keep_links=False, keep_vba=False):
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()
    if suffix in ('.xlsx', '.xlsm'):
        wb = openpyxl.load_workbook(file_path,
                                    keep_links=keep_links,
                                    keep_vba=keep_vba)
    elif suffix == '.xls':
        wb = convert_xls_to_xlsx(file_path)
    elif suffix == '.csv':
        wb = convert_csv_to_xlsx(file_path)
    else:
        return None
    return wb


def parse_range_address(address):
    """ 解析单元格范围地址。

    :param str address: 单元格范围地址，例如 'A1', 'A1:B3', '1:3', 'A:B' 等。
    :return dict: 一个包含 'left', 'top', 'right', 'bottom' 的字典。
    """
    # 初始化默认值
    left, right, top, bottom = None, None, None, None

    # 分割地址以获取开始和结束
    parts = address.split(":")
    start_cell = parts[0]
    end_cell = parts[1] if len(parts) > 1 else start_cell

    # 如果 start_cell 是行号
    if start_cell.isdigit():
        top = int(start_cell)
    else:
        # 尝试从 start_cell 提取列
        try:
            left = column_index_from_string(start_cell.rstrip('1234567890'))
            top = int(''.join(filter(str.isdigit, start_cell))) if any(
                char.isdigit() for char in start_cell) else None
        except ValueError:
            left = None

    # 如果 end_cell 是行号
    if end_cell.isdigit():
        bottom = int(end_cell)
    else:
        # 尝试从 end_cell 提取列
        try:
            right = column_index_from_string(end_cell.rstrip('1234567890'))
            bottom = int(''.join(filter(str.isdigit, end_cell))) if any(char.isdigit() for char in end_cell) else None
        except ValueError:
            right = None

    # 如果只提供了一个部分 (例如 '1', 'A')，将最大值设置为最小值
    if len(parts) == 1:
        right = left if left is not None else right
        bottom = top if top is not None else bottom

    return {"left": left, "top": top, "right": right, "bottom": bottom}


def get_addr_area(addr):
    """ 一个range描述的面积 """
    if ':' in addr:
        d = parse_range_address(addr)
        return (d['right'] - d['left'] + 1) * (d['bottom'] - d['top'] + 1)
    else:
        return 1


def build_range_address(left=None, top=None, right=None, bottom=None):
    """ 构建单元格范围地址。

    :return str: 单元格范围地址，例如 'A1', 'A1:B3', '1:3', 'A:B' 等。
    """
    start_cell = f"{get_column_letter(left) if left else ''}{top if top else ''}"
    end_cell = f"{get_column_letter(right) if right else ''}{bottom if bottom else ''}"

    # 当开始和结束单元格相同时，只返回一个单元格地址
    if start_cell == end_cell:
        return start_cell
    # 当其中一个单元格是空字符串时，只返回另一个单元格地址
    elif not start_cell or not end_cell:
        return start_cell or end_cell
    else:
        return f"{start_cell}:{end_cell}"


def combine_addresses(*addrs):
    # 初始化最小和最大行列值
    min_left, min_top, max_right, max_bottom = float('inf'), float('inf'), 0, 0

    # 遍历所有地址
    for addr in addrs:
        # 解析每个地址
        addr_dict = parse_range_address(addr)

        # 更新最小和最大行列值
        if addr_dict['left'] is not None:
            min_left = min(min_left, addr_dict['left'])
            max_right = max(max_right, addr_dict['right'] if addr_dict['right'] is not None else addr_dict['left'])
        if addr_dict['top'] is not None:
            min_top = min(min_top, addr_dict['top'])
            max_bottom = max(max_bottom, addr_dict['bottom'] if addr_dict['bottom'] is not None else addr_dict['top'])

    # 构建新的地址字符串
    new_addr = f"{get_column_letter(min_left)}{min_top}:{get_column_letter(max_right)}{max_bottom}"
    return new_addr


def is_string_type(value):
    """ 检查值是否为字符串类型，不是数值或日期类型 """
    # 首先检查日期类型
    try:
        pd.to_datetime(value, errors='raise')
        return False
    except (ValueError, TypeError, OverflowError):
        pass

    # 检查是否为浮点数类型
    try:
        float(value)
        return False
    except (ValueError, TypeError):
        return True


def __2_openpyxl_class():
    """ 对openpyxl已有类的功能的增强 """


class XlCell(openpyxl.cell.cell.Cell):  # 适用于 openpyxl.cell.cell.MergedCell，但这里不能多重继承

    def in_range(self):
        """ 判断一个单元格所在的合并单元格

        >> ws['C1'].in_range()
        <openpyxl.worksheet.cell_range.CellRange> A1:D3
        """
        ws = self.parent
        for rng in ws.merged_cells.ranges:
            if self.coordinate in rng:
                break
        else:  # 如果找不到则返回原值
            rng = self
        return rng

    def mcell(self):
        """返回“有效单元格”，即如果输入的是一个合并单元格，会返回该合并单元格左上角的单元格
        修改左上角单元格的值才是可行、有意义的

        因为跟合并单元格有关，所以 以m前缀 merge
        """
        if isinstance(self, MergedCell):
            ws = self.parent
            x, y = self.in_range().top[0]
            return ws.cell(x, y)
        else:
            return self

    def celltype(self, *, return_mode=False):
        """
        :param return_mode: 是否返回运算的中间结果信息
            主要是在type=2的情景，有时候需要使用rng变量，可以这里直接返回，避免外部重复计算
        :return: 单元格类型
            0：普通单元格
            1：合并单元格其他衍生位置
            2：合并单元格的左上角的位置

        TODO 这个函数还是可以看看能不能有更好的实现、提速
        """
        _type, status = 0, {}
        if isinstance(self, MergedCell):
            _type = 1
        elif isinstance(self.offset(1, 0), MergedCell) or isinstance(self.offset(0, 1), MergedCell):
            # 这里只能判断可能是合并单元格，具体是不是合并单元格，还要
            rng = self.in_range()
            status['rng'] = rng
            _type = 2 if hasattr(rng, 'size') else 0

        if return_mode:
            return _type, status
        else:
            return _type

    def isnone(self):
        """ 是普通单元格且值为None

        注意合并单元格的衍生单元格不为None
        """
        celltype = self.celltype()
        return celltype == 0 and self.value is None

    def clear(self):
        """ 清除数值、格式、合并单元格

        注意，如果self是合并单元格，分两种清空
            母格（左上角），会撤销合并到和母格数值、格式
            衍生格，只会撤销合并单元格，但不会清除母格的数值、格式

        :return: 涉及到合并单元格的情况，新单元格和原单元格已经不一样了，需要重新获取对象
        """
        ct, mid_result = self.celltype(return_mid_result=True)
        x = self
        if ct:  # 如果是合并单元格，取消该区域的合并单元格
            rng = mid_result['rng'] if ('rng' in mid_result) else self.in_range()
            self.parent.unmerge_cells(rng.coord)
            x = self.parent[self.coordinate]
        x.value = None
        x.style = 'Normal'
        return x

    def copy_cell_format(self, dst_cell):
        """ 单元格全格式复制，需要事先指定好新旧单元格的物理位置
        参考：https://stackoverflow.com/questions/23332259/copy-cell-style-openpyxl
        """
        from copy import copy
        if self.has_style:
            dst_cell.font = copy(self.font)  # 字体
            dst_cell.border = copy(self.border)  # 表格线
            dst_cell.fill = copy(self.fill)  # 填充色
            dst_cell.number_format = copy(self.number_format)  # 数字格式
            dst_cell.protection = copy(self.protection)  # 保护？
            dst_cell.alignment = copy(self.alignment)  # 对齐格式
            # dst_cell.style = self.style
        # if self.comment:
        # 这个会引发AttributeError。。。
        #       vml = fromstring(self.workbook.vba_archive.read(ws.legacy_drawing))
        #   AttributeError: 'NoneType' object has no attribute 'read'
        # dst_cell.comment = copy(cell.comment)
        # 就算开了keep_vba可以强制写入了，打开的时候文件可能还是会错

    def copy_cell(self, dst_cell):
        """ 单元格全格式、包括值的整体复制

        注意合并单元格比较复杂，比如要把 'A1:C3' 复制到 'A2:D4'，是会出现问题的
            在预先清空A2:D4数据的时候，会把
            一般这种清空，推荐先将数据库复制到一个临时sheet，再复制回原sheet更安全
        """
        from itertools import product
        ct, mid_result = self.celltype(return_mid_result=True)

        if ct == 0:  # 普通单元格，只复制值和格式
            dst_cell = dst_cell.clear()
            dst_cell.value = self.value
            self.copy_cell_format(dst_cell)
        elif ct == 2:  # 合并单元格，除了值和格式，要考虑单元格整体性的复制替换
            dst_cell = dst_cell.clear()
            rng = mid_result['rng'] if ('rng' in mid_result) else self.in_range()  # CellRange类型
            n, m = rng.size['rows'], rng.size['columns']  # 几行几列
            # 先把目标位置里的区域清空
            ws2 = dst_cell.parent
            x2, y2 = dst_cell.row, dst_cell.column
            for i, j in product(range(n), range(m)):
                ws2.cell(x2 + i, y2 + j).clear()
            # 拷贝数据
            dst_cell.value = self.value
            self.copy_cell_format(dst_cell)
            ws2.merge_cells(start_row=x2, start_column=y2, end_row=x2 + n - 1, end_column=y2 + m - 1)
        else:  # 合并单元格的衍生单元格复制时，不做任何处理
            return

    def down(self, count=1):
        """ 输入一个单元格，向下移动一格
        注意其跟offset的区别，如果cell是合并单元格，会跳过自身的衍生单元格

        :param count: 重复操作次数

        注意这里移动跟excel中操作也不太一样，设计的更加"原子化"，可以多配合cell.mcell功能使用。
        详见：【腾讯文档】cell移动机制说明 https://docs.qq.com/doc/DUkRUaFhlb3l4UG1P
        """

        def _func(cell):
            r, c = cell.row, cell.column
            if cell.celltype():
                rng = cell.in_range()
                r = rng.max_row
            return cell.parent.cell(r + 1, c)

        cell = self
        for _ in range(count):
            cell = _func(cell)
        return cell

    def right(self, count=1):
        def _func(cell):
            r, c = cell.row, cell.column
            if cell.celltype():
                rng = cell.in_range()
                c = rng.max_col
            return cell.parent.cell(r, c + 1)

        cell = self
        for _ in range(count):
            cell = _func(cell)
        return cell

    def up(self, count=1):
        def _func(cell):
            r, c = cell.row, cell.column
            if cell.celltype():
                rng = cell.in_range()
                r = rng.min_row
            return cell.parent.cell(max(r - 1, 1), c)

        cell = self
        for _ in range(count):
            cell = _func(cell)
        return cell

    def left(self, count=1):
        def _func(cell):
            r, c = cell.row, cell.column
            if cell.celltype():
                rng = cell.in_range()
                r = rng.min_row
            return cell.parent.cell(r, max(c - 1, 1))

        cell = self
        for _ in range(count):
            cell = _func(cell)
        return cell

    def fill_color(self, color, fill_type="solid", **kwargs):
        """ 封装一些我自己常用的填色方案 """
        from openpyxl.styles import PatternFill
        from pyxllib.cv.rgbfmt import RgbFormatter
        if isinstance(color, str):
            color = RgbFormatter.from_name(color)
        elif isinstance(color, (list, tuple)):
            color = RgbFormatter(*color)
        self.fill = PatternFill(fgColor=color.to_hex()[1:], fill_type=fill_type, **kwargs)

    def set_rich_value(self, value, color=None):
        """ 因为我经常文本和单元格背景色一起设置，所以这里封装一个接口
        如果只是普通的设置value，用cell.value就行，这个函数主要是设置富文本
        """
        self.value = value
        if color:
            self.fill_color(color)
        # todo 可以考虑扩展更多富文本样式，在这里统一设置

    def get_number_format(self):
        """ 相比源生的接口，有做了一些细节优化 """
        fmt = self.number_format
        # openpyxl的机制，如果没有配置日期格式，读取到的是默认的'mm-dd-yy'，其实在中文场景，默认格式应该是后者
        if fmt == 'mm-dd-yy':
            return 'yyyy/m/d'  # 中文的默认日期格式
        elif fmt == 'yyyy\-mm\-dd':  # 不知道为什么会有提取到这种\的情况，先暴力替换了
            fmt = 'yyyy-mm-dd'
        return fmt

    def get_render_value(self):
        """ 得到单元格简单渲染后的效果
        py里不可能对excel的所有格式进行全覆盖，只是对场景格式进行处理

        注意，遇到公式是很难计算处理的，大概率只能保持原公式显示
        因为日期用的比较多，需要时常获得真实的渲染效果，所以这里封装一个接口
        """
        x = self.value
        xl_fmt = self.get_number_format()
        return xl_render_value(x, xl_fmt)

    def address(self):
        if isinstance(self, openpyxl.cell.cell.Cell):
            return self.coordinate
        else:  # 否则认为是合并单元格
            return str(self)  # 标准库有自带方法，可以直接转的


# 只有cell和mergecell都共同没有的成员方法，才添加进去
__members = set(dir(XlCell)) - set(dir(openpyxl.cell.cell.Cell)) - \
            set(dir(openpyxl.cell.cell.MergedCell)) - {'__dict__'}
inject_members(XlCell, openpyxl.cell.cell.Cell, __members)
inject_members(XlCell, openpyxl.cell.cell.MergedCell, __members)


class XlWorksheet(openpyxl.worksheet.worksheet.Worksheet):
    """ 扩展标准的Workshhet功能 """

    def get_raw_usedrange(self):
        raw_used_range = build_range_address(left=self.min_column, top=self.min_row,
                                             right=self.max_column, bottom=self.max_row)
        return raw_used_range

    def is_empty_row(self, row, start_col, end_col):
        if not hasattr(self, 'is_empty_row_cache'):
            self.is_empty_row_cache = {}
        key = (row, start_col, end_col)

        def is_empty_row_core():
            cur_col = start_col
            # 特地提前检查下最后一列的那个单元格
            if self.cell(row, end_col).value is not None:
                return False
            while cur_col <= end_col:
                if self.cell(row, cur_col).value is not None:
                    return False
                # 步长随着尝试的增加，也逐渐降低采样率
                n = cur_col - start_col + 1
                # 在最大值m=16384列情况下，/1000，最多检索3404个单元格，/100，最多检索569次，/50最多检索320次
                # cur_col += (n // 50) + 1
                # 再变形，加强前面权重，大大降低后面权重
                if n <= 100:
                    cur_col += 1
                else:  # 最多54次
                    cur_col += (n // 10)

            return True

        if key not in self.is_empty_row_cache:
            self.is_empty_row_cache[key] = is_empty_row_core()

        return self.is_empty_row_cache[key]

    def is_empty_column(self, col, start_row, end_row):
        if not hasattr(self, 'is_empty_column_cache'):
            self.is_empty_column_cache = {}
        key = (col, start_row, end_row)

        def is_empty_column_core():
            cur_row = start_row
            # 特地提前检查下最后一行的那个单元格
            if self.cell(end_row, col).value is not None:
                return False
            while cur_row <= end_row:
                if self.cell(cur_row, col).value is not None:
                    return False
                n = cur_row - start_row + 1
                # 在最大值n=1048576行情况下，/1000，最多检索7535个单元格，/100，最多检索987次，/50最多检索530次
                cur_row += (n // 1000) + 1
            return True

        if key not in self.is_empty_column_cache:
            self.is_empty_column_cache[key] = is_empty_column_core()

        return self.is_empty_column_cache[key]

    def find_last_non_empty_row(self, start_row, end_row, start_col, end_col, m=30):
        # 1 如果剩余行数不多（小于等于m），直接遍历这些行
        if end_row - start_row <= m:  # 这里是兼容start_row大于end_row的情况的
            for row in range(end_row, start_row - 1, -1):
                if not self.is_empty_row(row, start_col, end_col):
                    return row
            return -1  # 如果都是空的，则返回-1

        # 2 计算分割点
        intervals = [(end_row - start_row) // (m - 1) * i + start_row for i in range(m - 1)] + [end_row]

        # 3 反向遍历这些分割点，找到第一个非空行
        for i in reversed(range(len(intervals))):
            # 检查点全部都空，也不能判定全空，要检查前面半个区间
            if i == 0 or not self.is_empty_row(intervals[i], start_col, end_col):
                # 如果这是最后一个分割点，则直接返回它
                if i == m - 1:
                    return intervals[i]
                # 否则，在这个分割点和下一个（一半二分的位置）之间递归查找
                return self.find_last_non_empty_row(intervals[i],
                                                    intervals[min(i + m // 2, m - 1)],
                                                    start_col,
                                                    end_col,
                                                    m + 1)

        # 如果所有分割点都是空的，则返回-1
        return -1

    def find_last_non_empty_column(self, start_col, end_col, start_row, end_row, m=30):
        # dprint(end_col)

        # 1 如果剩余列数不多（小于等于m），直接遍历这些列
        if end_col - start_col <= m:
            for col in range(end_col, start_col - 1, -1):
                if not self.is_empty_column(col, start_row, end_row):
                    return col
            return -1  # 如果都是空的，则返回-1

        # 2 计算分割点
        intervals = [(end_col - start_col) // (m - 1) * i + start_col for i in range(m - 1)] + [end_col]

        # 3 反向遍历这些分割点，找到第一个非空列
        for i in reversed(range(len(intervals))):
            if i == 0 or not self.is_empty_column(intervals[i], start_row, end_row):
                # 如果这是最后一个分割点，则直接返回它
                if i == m - 1:
                    return intervals[i]
                # 否则，在这个分割点和下一个（一半二分的位置）之间递归查找
                return self.find_last_non_empty_column(intervals[i],
                                                       intervals[min(i + m // 2, m - 1)],
                                                       start_row,
                                                       end_row,
                                                       m + 1)
        # 如果所有分割点都是空的，则返回-1
        return -1

    def find_first_non_empty_row(self, start_row, end_row, start_col, end_col, m=30):
        # 1 如果剩余行数不多（小于等于m），直接遍历这些行
        if end_row - start_row <= m:
            for row in range(start_row, end_row + 1):
                if not self.is_empty_row(row, start_col, end_col):
                    return row
            return -1  # 如果都是空的，则返回-1

        # 2 计算分割点
        intervals = [(end_row - start_row) // (m - 1) * i + start_row for i in range(m - 1)] + [end_row]

        # 3 正向遍历这些分割点，找到第一个非空行
        for i in range(len(intervals)):
            if i == m - 1 or not self.is_empty_row(intervals[i], start_col, end_col):
                # 如果这是第一个分割点，则直接返回它
                if i == 0:
                    return intervals[i]
                # 否则，在这个分割点和前一个（一半二分的位置）之间递归查找
                return self.find_first_non_empty_row(intervals[max(i - m // 2, 0)],
                                                     intervals[i],
                                                     start_col,
                                                     end_col,
                                                     m + 1)
        # 如果所有分割点都是空的，则返回-1
        return -1

    def find_first_non_empty_column(self, start_col, end_col, start_row, end_row, m=30):
        # 1 如果剩余列数不多（小于等于m），直接遍历这些列
        if end_col - start_col <= m:
            for col in range(start_col, end_col + 1):
                if not self.is_empty_column(col, start_row, end_row):
                    return col
            return -1  # 如果都是空的，则返回-1

        # 2 计算分割点
        intervals = [(end_col - start_col) // (m - 1) * i + start_col for i in range(m - 1)] + [end_col]

        # 3 正向遍历这些分割点，找到第一个非空列
        for i in range(len(intervals)):
            if i == m - 1 or not self.is_empty_column(intervals[i], start_row, end_row):
                # 如果这是第一个分割点，则直接返回它
                if i == 0:
                    return intervals[i]
                # 否则，在这个分割点和前一个（一半二分的位置）之间递归查找
                return self.find_first_non_empty_column(intervals[max(i - m // 2, 0)],
                                                        intervals[i],
                                                        start_row,
                                                        end_row,
                                                        m + 1)
        # 如果所有分割点都是空的，则返回-1
        return -1

    def get_usedrange(self):
        """ 定位有效数据区间。

        背景：
            在Excel工作表中，经常需要确定包含数据的有效区域。这是因为工作表可能包含大量的空白区域，
            而实际数据仅占据一部分空间。有效地识别这个区域对于进一步的数据处理和分析至关重要。
            如果暴力一行行遍历，遇到XFD1048576这种覆盖全范围的表，运行肯定会超时。

        求解思路：
            为了高效地定位有效数据区间，我们采用了分割点技术和二分查找的思想。
            具体解释，以找第1~100行中最后一个非空行为例。
            可以判断第1、50、100行是否是空行，如果50、100都是空行，那空行应该在第1~50的范围里，然后再判定第25行是否为空行。
            但二分法可能不严谨，第50、100都是空行，中间也有可能有内容，比如第1~80行本来其实都有内容，只是恰好第50行空了。
            所以在二分法的基础上，还需要把1~100行等间距取m个采样点来辅助检查。

            通过调m的值的变化方法，可以在速度和精度之间做一个权衡。
            这四个定界符，最最慢的是find_last_non_empty_row，重点调这个。

        四个边界判定函数，工程上是可以整合的，但是整合后，太多if分支操作，会降低效率，这里为了运行速度，就拆分4个实现更好。
            具体实现中，还有其他一些细节考虑优化。
            比如先找最后行，再最后列，再第一行，第一列，这个顺序是有讲究的，可以减少很多不必要的遍历。
            因为数据一般是比较高和窄的，所以应该先对行做处理。以及前面出现空区域的概率小，可以后面再处理。
            而且就openpyxl而言，对列的遍历也是远慢于行的遍历的。

        :param reset_bounds: 计算出新区域后，是否重置ws的边界值
        """
        if not hasattr(self, 'usedrange_cache'):
            # 初始化边界值
            left, right, top, bottom = self.min_column, self.max_column, self.min_row, self.max_row

            # start_time = time.time()
            # 使用优化后的函数找到最下方的行和最右边的列
            bottom = self.find_last_non_empty_row(top, bottom, left, right)
            if bottom == -1:
                return 'A1'  # 空表返回A1占位
            right = self.find_last_non_empty_column(left, right, top, bottom)
            if right == -1:
                return 'A1'

            # 使用优化后的函数找到最上方的行和最左边的列
            top = self.find_first_non_empty_row(top, bottom, left, right)
            if top == -1:
                return 'A1'
            left = self.find_first_non_empty_column(left, right, top, bottom)
            if left == -1:
                return 'A1'
            # get_global_var('get_usedrange_time')[-1] += time.time() - start_time

            # 2 然后还要再扩范围（根据合并单元格情况）
            # start_time = time.time()
            top0, bottom0, left0, right0 = top, bottom, left, right
            for merged_range in self.merged_cells.ranges:
                l, t, r, b = merged_range.bounds
                if top0 <= b <= bottom0 or top0 <= t <= bottom0:
                    if left0 <= r and l < left:
                        left = l
                    if l <= right0 and r > right:
                        right = r
                if left0 <= r <= right0 or left0 <= l <= right0:
                    if top0 <= b and t < top:
                        top = t
                    if t <= bottom0 and b > bottom:
                        bottom = b
            # get_global_var('expandrange_time')[-1] += time.time() - start_time

            self.used_range = build_range_address(left=left, top=top, right=right, bottom=bottom)

        return self.used_range

    def copy_worksheet(self, dst_ws):
        """跨工作薄时复制表格内容的功能
        openpyxl自带的Workbook.copy_worksheet没法跨工作薄复制，很坑

        src_ws.copy_worksheet(dst_ws)
        """
        # 1 取每个单元格的值
        for row in self:
            for cell in row:
                try:
                    cell.copy_cell(dst_ws[cell.coordinate])
                except AttributeError:
                    pass
        # 2 合并单元格的处理
        for rng in self.merged_cells.ranges:
            dst_ws.merge_cells(rng.ref)
        # 3 其他表格属性的复制
        # 这个从excel读取过来的时候，是不准的，例如D3可能因为关闭时停留窗口的原因误跑到D103
        # dprint(origin_ws.freeze_panes)
        # target_ws.freeze_panes = origin_ws.freeze_panes

    def _cells_by_row(self, min_col, min_row, max_col, max_row, values_only=False):
        """ openpyxl的这个迭代器，遇到合并单元格会有bug
        所以我把它重新设计一下~~
        """
        for row in range(min_row, max_row + 1):
            cells = (self.cell(row=row, column=column) for column in range(min_col, max_col + 1))
            if values_only:
                # yield tuple(cell.value for cell in cells)  # 原代码
                yield tuple(getattr(cell, 'value', None) for cell in cells)
            else:
                yield tuple(cells)

    def search(self, pattern, min_row=None, max_row=None, min_col=None, max_col=None, order=None, direction=0):
        """ 查找满足pattern正则表达式的单元格

        :param pattern: 正则匹配式，可以输入re.complier对象
            会将每个单元格的值转成str，然后进行字符串规则search匹配
                注意日期的本质是一个数字，pattern支持输入一个datetime.date类型，会自动转为excel的日期值
            支持多层嵌套 ['模块一', '属性1']
        :param direction: 只有在 pattern 为数组的时候有用
            pattern有多组时，会嵌套找单元格
            每计算出一个条件后，默认取该单元格下方的子区间 axis=0
            如果不是下方，而是右方，可以设置为1
            还有奇葩的上方、左方，可以分别设置为2、3
        :param order: 默认None，也就是 [1, 2] 的效果，规律详见product接口

        >> wb = openpyxl.load_workbook(filename='2020寒假教材各地区数量统计最新2020.1.1.xlsx')
        >> ws = Worksheet(wb['预算总表'])
        >> ws.search('年段')
        <Cell '预算总表'.B2>
        """
        if not hasattr(self, 'search_cache'):
            self.search_cache = {}

        if isinstance(pattern, list):
            pattern = tuple(pattern)

        key = (pattern, min_row, max_row, min_col, max_col, order, direction)

        def get_search_core():
            nonlocal pattern
            # 1 定界
            x1, x2 = max(min_row or 1, 1), min(max_row or self.max_row, self.max_row)
            y1, y2 = max(min_col or 1, 1), min(max_col or self.max_column, self.max_column)

            # 2 遍历
            if isinstance(pattern, datetime.date):
                pattern = f'^{(pattern - datetime.date(1899, 12, 30)).days}$'

            if isinstance(pattern, tuple):
                cel = None
                for p in pattern:
                    cel = self.search(p, x1, x2, y1, y2, order)
                    if cel:
                        # up, down, left, right 找到的单元格四边界
                        l, u, r, d = getattr(cel.in_range(), 'bounds', (cel.column, cel.row, cel.column, cel.row))
                        if direction == 0:
                            x1, x2, y1, y2 = max(x1, d + 1), x2, max(y1, l), min(y2, r)
                        elif direction == 1:
                            x1, x2, y1, y2 = max(x1, u), min(x2, d), max(y1, r + 1), y2
                        elif direction == 2:
                            x1, x2, y1, y2 = x1, min(x2, u - 1), max(y1, l), min(y2, r)
                        elif direction == 3:
                            x1, x2, y1, y2 = max(x1, u), min(x2, d), y1, min(y2, l - 1)
                        else:
                            raise ValueError(f'direction参数值错误{direction}')
                    else:
                        return None
                return cel
            else:
                if isinstance(pattern, str): pattern = re.compile(pattern)
                for x, y in product(range(x1, x2 + 1), range(y1, y2 + 1), order=order):
                    cell = self.cell(x, y)
                    if cell.celltype() == 1: continue  # 过滤掉合并单元格位置
                    if pattern.search(str(cell.value)): return cell  # 返回满足条件的第一个值

        if key not in self.search_cache:
            self.search_cache[key] = get_search_core()

        return self.search_cache[key]

    findcel = search

    def findrow(self, pattern, *args, **kwargs):
        cel = self.findcel(pattern, *args, **kwargs)
        return cel.row if cel else 0

    def findcol(self, pattern, *args, **kwargs):
        cel = self.findcel(pattern, *args, **kwargs)
        return cel.column if cel else 0

    def browser(self):
        """注意，这里会去除掉合并单元格"""
        browser(pd.DataFrame(self.values))

    def select_columns(self, columns, column_name='searchkey'):
        r""" 获取表中columns属性列的值，返回dataframe数据类型

        :param columns: 搜索列名使用正则re.search字符串匹配查找
            可以单列：'attr1'，找到列头后，会一直往后取到最后一个非空值
            也可以多列： ['attr1', 'attr2', 'attr3']
                会结合多个列标题定位，数据从最大的起始行号开始取，
                （TODO 截止到最末非空值所在行  未实现，先用openpyxl自带的max_row判断，不过这个有时会判断过大）
            遇到合并单元格，会寻找其母单元格的值填充
        :param column_name: 返回的df。列名
            origin，原始的列名
            searchkey，搜索时用的查找名
        """
        if not isinstance(columns, (list, tuple)):
            columns = [columns]

        # 1 找到所有标题位置，定位起始行
        cels, names, start_line = [], [], -1
        for search_name in columns:
            cel = self.findcel(search_name)
            if cel:
                cels.append(cel)
                if column_name == 'searchkey':
                    names.append(str(search_name))
                elif column_name == 'origin':
                    if isinstance(search_name, (list, tuple)) and len(search_name) > 1:
                        names.append('/'.join(list(search_name[:-1]) + [str(cel.value)]))
                    else:
                        names.append(str(cel.value))
                else:
                    raise ValueError(f'{column_name}')
                start_line = max(start_line, cel.down().row)
            else:
                dprint(search_name)  # 找不到指定列

        # 2 获得每列的数据
        datas = {}
        for k, cel in enumerate(cels):
            if cel:
                col = cel.column
                li = []
                for i in range(start_line, self.max_row + 1):
                    v = self.cell(i, col).mcell().value  # 注意合并单元格的取值
                    li.append(v)
                datas[names[k]] = li
            else:
                # 如果没找到列，设一个空列
                datas[names[k]] = [None] * (self.max_row + 1 - start_line)
        df = pd.DataFrame(datas)

        # 3 去除所有空行数据
        df.dropna(how='all', inplace=True)

        return df

    def copy_range(self, src_addr, dst_cell, *, temp_sheet=False, return_mid_result=False):
        """ 将自身cell_range区间的内容、格式，拷贝到目标dst_cell里

        :param str src_addr: 自身的一片单元格范围
            支持输入格式：str --> cell
            支持范式：普通单元格 --> 合并单元格
        :param dst_cell: 要复制到的目标单元格位置
            输入起点、单个位置
            一般只有同个工作表ws要考虑赋值顺序问题，防止引用前已被更新覆盖
                但有个极端情况：循环引用的公式计算，在跨ws的时候如果不考虑顺序也可能出错，但这种情况太复杂的，这里不可能去处理
        :param temp_sheet: 当拷贝中涉及合并单元格等重叠位置问题时，建议开启该参数，用中间数据表过渡下

        这个算法主要难点，是要考虑合并单元格的情况比较复杂。否则本身逻辑并不难。

        >> ws1.copy_range('A1:C3', ws2.cell('C2'))  # 将ws1的A1:C3数据复制到ws2的C2里
        """
        from itertools import product

        # 0 中间表
        mid_result = {}
        if temp_sheet:
            ws3 = self.parent.create_sheet('__copy_range')
            mid_result = self.copy_range(src_addr, ws3['A1'], return_mid_result=True)
            ws1 = ws3
            src_addr = f'A1:{excel_addr(mid_result["n"], mid_result["m"])}'
        else:
            ws1 = self
        ws2 = dst_cell.parent

        # 1 坐标计算
        # 用ws1[]比用CellRange更精准，还能处理"1:3"这种泛式定位，会根据max_column智能判定边界单元格
        src_cells = ws1[src_addr]
        # 强制转为n*m的二维tuple数组结构
        if not isinstance(src_cells, tuple):
            src_cells = (src_cells,)
        if not isinstance(src_cells[0], tuple):
            src_cells = (src_cells,)
        # 关键信息
        n, m = len(src_cells), len(src_cells[0])  # 待复制的数据是几行几列
        src_cell = src_cells[0][0]
        x1, y1 = src_cell.row, src_cell.column  # ws1数据起始位置，x是行，y是列
        x2, y2 = dst_cell.row, dst_cell.column
        bias_rows, bias_cols = x2 - x1, y2 - y1
        mid_result['n'] = n
        mid_result['m'] = m

        # 2 将src内容拷贝过去
        # 注意拷贝顺序跟移动方向是有关系的，要防止被误覆盖，复制了新的值，而非原始值
        r = sorted(range(n), reverse=bias_rows > 0)  # 要拷贝的每行
        c = sorted(range(m), reverse=bias_cols > 0)
        for i, j in product(r, c):  # openpyxl好像没有复制range的功能？
            ws1.cell(x1 + i, y1 + j).copy_cell(ws2.cell(x2 + i, y2 + j))

        # 3 收尾
        if temp_sheet:
            self.parent.remove(ws1)

        if return_mid_result:
            return mid_result

    def reindex_columns(self, orders):
        """ 重新排列表格的列顺序

        >> ws.reindex_columns('I,J,A,,,G,B,C,D,F,E,H,,,K'.split(','))

        TODO 支持含合并单元格的整体移动？
        """
        from openpyxl.utils.cell import column_index_from_string
        max_row, max_column = self.max_row, self.max_column
        for j, col in enumerate(orders, 1):
            if not col: continue
            self.copy_range(f'{col}1:{col}{max_row}', self[excel_addr(1, max_column + j)])
        self.delete_cols(1, max_column)

    def to_html(self, *, border=1,
                style='border-collapse:collapse; text-indent:0; margin:0;') -> str:
        r"""
        .from_latex(r'''\begin{tabular}{|c|c|c|c|}
                      \hline
                      1 & 2 & & 4\\
                      \hline
                      \end{tabular}''').to_html())

        ==>

        <table border="1" style="border-collapse: collapse;">
          <tr>
            <td style="text-align:center">
              1
            </td>
            <td style="text-align:center">
              2
            </td>
            <td style="text-align:center"></td>
            <td style="text-align:center">
              4
            </td>
          </tr>
        </table>
        """
        from yattag import Doc

        doc, tag, text = Doc().tagtext()
        tag_attrs = [('border', border), ('style', style)]
        # if self.data_tex:  # 原来做latex的时候有的一个属性
        #     tag_attrs.append(('data-tex', self.data_tex))

        ws = self
        with tag('table', *tag_attrs):
            # dprint(ws.max_row, ws.max_column)
            cols = ws.max_column
            for i in range(1, ws.max_row + 1):
                # TODO 这样跳过其实也不太好，有时候可能就是想创建一个空内容的表格
                for j in range(1, cols + 1):
                    if not ws.cell(i, j).isnone():
                        break
                else:  # 如果没有内容，跳过该行
                    continue

                with tag('tr'):
                    for j in range(1, cols + 1):
                        # ① 判断单元格类型
                        cell = ws.cell(i, j)
                        celltype = cell.celltype()
                        if celltype == 1:  # 合并单元格的衍生单元格
                            continue
                        elif cell.isnone():  # 其他正常的空单元格
                            with tag('td'):
                                doc.asis('')
                            continue
                        # ② 对齐等格式控制
                        params = {}
                        if celltype == 2:  # 合并单元格的左上角
                            rng = cell.in_range()
                            params['rowspan'] = rng.size['rows']
                            params['colspan'] = rng.size['columns']
                        if cell.alignment.horizontal:
                            params['style'] = 'text-align:' + cell.alignment.horizontal
                        # if cell.alignment.vertical:
                        #     params['valign'] = cell.alignment.vertical
                        with tag('td', **params):
                            v = str(cell.value)
                            # if not v: v = '&nbsp;'  # 200424周五15:40，空值直接上平台表格会被折叠，就加了个空格
                            doc.asis(v)  # 不能用text，html特殊字符不用逃逸
        # res = indent(doc.getvalue(), indent_text=True)  # 美化输出模式。但是这句在某些场景会有bug。
        res = doc.getvalue()  # 紧凑模式

        return res

    def init_from_latex(self, content):
        """ 注意没有取名为from_latex，因为ws是事先创建好的，这里只是能输入latex代码进行初始化而已 """
        from openpyxl.styles import Border, Alignment, Side

        from pyxllib.text.pupil import grp_bracket
        from pyxllib.text.latex import TexTabular

        BRACE2 = grp_bracket(2, inner=True)
        BRACE5 = grp_bracket(5, inner=True)

        # 暂时统一边框线的样式 borders。不做细化解析
        double = Side(border_style='thin', color='000000')

        ws = self

        # 处理表头
        data_tex = re.search(r'\\begin{tabular}\s*(?:\[.*?\])?\s*' + BRACE5, content).group(1)
        col_pos = TexTabular.parse_align(data_tex)  # 每列的格式控制
        # dprint(self.data_tex, col_pos)
        total_col = len(col_pos)
        # 删除头尾标记
        s = re.sub(r'\\begin{tabular}(?:\[.*?\])?' + BRACE5, '', re.sub(r'\\end{tabular}', '', content))
        row, col = 1, 1

        # 先用简单不严谨的规则确定用全网格，还是无网格
        # if '\\hline' not in s and '\\midrule' not in s:
        #     border = 0

        # 用 \\ 分割处理每一行
        for line in re.split(r'\\\\(?!{)', s)[:-1]:
            # dprint(line)
            # 1 处理当前行的所有列元素
            cur_line = line
            # dprint(line)
            # 清除特殊格式数据
            cur_line = re.sub(r'\\cmidrule' + BRACE2, '', cur_line)
            cur_line = re.sub(r'\\cline' + BRACE2, '', cur_line)
            for t in (r'\midrule', r'\toprule', r'\bottomrule', r'\hline', '\n'):
                cur_line = cur_line.replace(t, '')

            # 遍历每列
            # dprint(cur_line)
            for item in cur_line.strip().split('&'):
                item = item.strip()
                # dprint(item)
                cur_loc = excel_addr(row, col)
                # dprint(row, col)

                if 'multicolumn' in item:
                    size, align, text = TexTabular.parse_multicolumn(item)
                    align = TexTabular.parse_align(align) if align else col_pos[col - 1]  # 如果没有写对齐，用默认列的格式
                    n, m = size
                    # 左右对齐，默认是left
                    align = {'l': 'left', 'c': 'center', 'r': 'right'}.get(align, 'left')
                    cell = ws[cur_loc].mcell()
                    if cell.value:
                        cell.value += '\n' + text
                    else:
                        cell.value = text
                    ws[cur_loc].alignment = Alignment(horizontal=align, vertical='center')
                    merge_loc = excel_addr(row + n - 1, col + m - 1)
                    ws.merge_cells(f'{cur_loc}:{merge_loc}')
                    col += m
                elif 'multirow' in item:
                    n, bigstructs, width, fixup, text = TexTabular.parse_multirow(item, brace_text_only=False)
                    try:
                        ws[cur_loc] = text
                    except AttributeError:
                        # 遇到合并单元格重叠问题，就修改旧的合并单元格，然后添加新单元格
                        # 例如原来 A1:A3 是一个合并单元格，现在要独立一个A3，则原来的部分重置为A1:A2
                        rng = ws[cur_loc].in_range()
                        ws.unmerge_cells(rng.coord)  # 解除旧的合并单元格
                        ws.merge_cells(re.sub(r'\d+$', f'{row - 1}', rng.coord))
                        ws[cur_loc] = text
                    align = {'l': 'left', 'c': 'center', 'r': 'right'}.get(col_pos[col - 1], 'left')
                    ws[cur_loc].alignment = Alignment(horizontal=align, vertical='center')
                    # dprint(item, row, n)
                    merge_loc = excel_addr(row + n - 1, col)
                    ws.merge_cells(f'{cur_loc}:{merge_loc}')
                    col += 1
                else:
                    if ws[cur_loc].celltype() == 0:
                        ws[cur_loc].value = item
                        # dprint(item, col_pos, col)
                        align = {'l': 'left', 'c': 'center', 'r': 'right'}.get(col_pos[col - 1], 'left')
                        ws[cur_loc].alignment = Alignment(horizontal=align, vertical='center')
                    col += 1

            # 2 其他border等格式控制
            if r'\midrule' in line or r'\toprule' in line or r'\bottomrule' in line or r'\hline' in line:
                # 该行画整条线
                loc_1 = excel_addr(row, 1)
                loc_2 = excel_addr(row, total_col)
                comb_loc = f'{loc_1}:{loc_2}'
                for cell in ws[comb_loc][0]:
                    cell.border = Border(top=double)
            if r'\cmidrule' in line:
                for match in re.findall(r'\\cmidrule{([0-9]+)-([0-9]+)}', line):
                    loc_1 = excel_addr(row, match[0])
                    loc_2 = excel_addr(row, match[1])
                    comb_loc = f'{loc_1}:{loc_2}'
                    for cell in ws[comb_loc][0]:
                        cell.border = Border(top=double)
            if r'\cline' in line:
                for match in re.findall(r'\\cline{([0-9]+)-([0-9]+)}', line):
                    loc_1 = excel_addr(row, match[0])
                    loc_2 = excel_addr(row, match[1])
                    comb_loc = f'{loc_1}:{loc_2}'
                    for cell in ws[comb_loc][0]:
                        cell.border = Border(top=double)
            row, col = row + 1, 1

    def to_latex(self):
        from pyxllib.text.latex import TexTabular

        ws = self

        li = []
        n, m = ws.max_row, ws.max_column
        format_count = [''] * m  # 记录每一列中各种对齐格式出现的次数
        merge_count = [0] * m  # 每列累积被合并行数，用于计算cline

        li.append('\\hline')
        for i in range(1, n + 1):
            if ws.cell(i, 1).isnone(): continue
            line = []
            j = 1
            while j < m + 1:
                cell = ws.cell(i, j)
                celltype = cell.celltype()
                if celltype == 0:  # 普通单元格
                    line.append(str(cell.value))
                elif celltype == 1:  # 合并单元格的衍生单元格
                    mcell = cell.mcell()  # 找出其母单元格
                    if mcell.column == cell.column:
                        columns = mcell.in_range().size['columns']
                        if columns > 1:
                            line.append(f'\\multicolumn{{{columns}}}{{|c|}}{{}}')  # 加一个空的multicolumn
                        else:
                            line.append('')  # 加一个空值
                elif celltype == 2:  # 合并单元格的左上角
                    rng = cell.in_range()
                    v = cell.value
                    rows, columns = rng.size['rows'], rng.size['columns']
                    if rows > 1:  # 有合并行
                        v = f'\\multirow{{{rows}}}*{{{v}}}'
                        for k in range(j, j + columns): merge_count[k - 1] = rows - 1
                    if columns > 1:  # 有合并列
                        # horizontal取值有情况有
                        # {‘center’, ‘centerContinuous’, ‘fill’, ‘left’, ‘justify’, ‘distributed’, ‘right’, ‘general’}
                        # 所以如果不是left、center、right，改成默认c
                        align = cell.alignment.horizontal[0]
                        if align not in 'lcr': align = 'c'
                        v = f'\\multicolumn{{{columns}}}{{|{align}|}}{{{v}}}'
                    line.append(str(v))
                    j += columns - 1
                if cell.alignment.horizontal:
                    format_count[j - 1] += cell.alignment.horizontal[0]
                j += 1
            li.append(' & '.join(line) + r'\\ ' + TexTabular.create_cline(merge_count))
            merge_count = [(x - 1 if x > 0 else x) for x in merge_count]
        li.append('\\end{tabular}\n')
        head = '\\begin{tabular}' + TexTabular.create_formats(format_count)
        li = [head] + li  # 开头其实可以最后加，在遍历中先确认每列用到最多的格式情况

        return '\n'.join(li)

    def cell2(self, row, column, value=None):
        """ 相比原版的cell，支持智能定位单元格位置

        :param row:
            int，第几行，从1开始编号
            dict, {k: v}，找k所在列，且值为v的行。结果可能不唯一，只会取第1个匹配项。
                支持多键同时检索
                目前v只用普通数值，以后可以扩展更灵活丰富的格式
        :param column: 用字段名column找单元格
            int，正常第几列，从1开始编号
            str,
                优先判断是否为纯大写字幕，使用'P'、'AZ'等进行解析
                其次使用findcol检索对应单元格
            List[str]，类似str，findcol
        """
        # 1 智能 row
        if isinstance(row, dict):
            idx_name = tuple(row.keys())[0]
            cols = {self.findcol(k): v for k, v in row.items()}
            for i in self.iterrows(idx_name):
                logo = True
                for c, v in cols.items():
                    if self.cell(i, c).value != v:
                        logo = False
                        break
                if logo:
                    row = i
                    break
            else:
                raise ValueError('Not find cell')

        # 2 智能 column
        if isinstance(column, int):
            pass
        elif isinstance(column, str) and re.match(r'[A-Z]+$', column):
            column = column_index_from_string(column)
        else:
            column = self.findcol(column)
            if not column:
                return None

        # 3 单元格
        # cell = self.cell(row, column, value)  # 这种写法好像有bug，写长文本的时候，后丢掉后半部分
        cell = self.cell(row, column)
        if value is not None:
            cell.value = value
        return cell

    def iterrows(self, key_column_name, mode='auto', *, to_dict=None):
        """ 通过某个属性列作为key，判断数据所在行

        正常遍历行用iterrows，离散找数据用cell2

        :param key_column_name: 参考的主要字段名，判断数据起始行
        :param mode: 计算数据范围的一些细分方法，目前主要是数据结束位置的判断方法
            default: 从ws.max_row往前找到第一个key非空的单元格
            auto: 基于一套智能的usedrange判定方法
            any_content: 从ws.max_row往前找到第一个含有值的行
            ... 待开发更多需求
        :param list[str] to_dict: 写出属性名，迭代的时候，返回除了下标，还有转换出的字典数据
        :return: 返回range类型，可以直接用于for循环
        """
        # 1 起始行
        cel = self.findcel(key_column_name).down()
        min_row = cel.row

        # 2 终止行
        max_row = self.max_row

        if mode == 'default':
            col = cel.column
            while max_row > min_row:
                if self.cell(max_row, col).value:
                    break
                max_row -= 1
        elif mode == 'any_content':
            max_column = self.max_column
            while max_row > min_row:
                empty_line = True
                for j in range(1, max_column + 1):
                    if self.cell(max_row, j).value:
                        empty_line = False
                        break
                if not empty_line:
                    break
        elif mode == 'auto':
            rng = parse_range_address(self.get_usedrange())
            max_row = rng['bottom']
        else:
            raise NotImplementedError(f'{mode}')

        if to_dict:
            data = []
            for i in range(min_row, max_row + 1):
                msg = {}
                for k in to_dict:
                    msg[k] = self.cell2(i, k).value
                data.append([i, msg])
            return data
        else:
            return range(min_row, max_row + 1)

    def find_head_data_range(self, ref_col_name):
        """ 查找表格的表头、数据所在区域

        可以把表格分成两大块：表头head，数据data
        每块数据都是一个矩形，有四个边界：ltrb

        :param ref_col_name: 参考的主列字段名字（如果是复合表头，需要表头最后一行的某个字段名）
            用这个名字才能区分出表头、数据划分界限在哪

        TODO right、bottom会多出一些空行、空列，怎么优化？
        """
        cel = self.findcel(ref_col_name)
        data_start_row = cel.down().row

        return {
            # 1 关键字段所在位置
            'cel': cel,
            'row': cel.row,
            'col': cel.column,
            # 2 表格左右区间
            'left': self.min_column,
            'right': self.max_column,
            # 3 表头和数据划分行
            'head_top': self.min_row,
            'head_bottom': data_start_row - 1,
            'data_top': data_start_row,
            'data_bottom': self.max_row,
        }

    def autofit(self):
        """ 自动调整工作表中所有列的宽度
        这里并不是excel自带标准的autofit，而是一种近似算法
        """

        def adjusted_column_width(cell_value):
            """
            根据单元格的内容调整列宽。
            假设中文字符的宽度是拉丁字符的两倍。
            """
            width_constant = 1.2  # 根据所需宽度调整此常数
            try:
                chinese_characters = sum(1 for char in cell_value if '\u4e00' <= char <= '\u9fff')
                latin_characters = len(cell_value) - chinese_characters
                return (chinese_characters * 2 + latin_characters) * width_constant
            except TypeError:
                return 10 * width_constant  # 如果单元格没有值或非字符串值则返回默认宽度

        for col in self.columns:
            max_width = 0
            column = [cell for cell in col]
            for cell in column:
                adjusted_width = adjusted_column_width(cell.value)
                if adjusted_width > max_width:
                    max_width = adjusted_width
            # 找到列中的第一个非合并单元格
            first_non_merged_cell = next((cell for cell in column if not isinstance(cell, MergedCell)), None)
            if first_non_merged_cell:
                self.column_dimensions[first_non_merged_cell.column_letter].width = min(max_width, 100)
            # 列宽最多设置到100，再大就增设自动换行来实现排版
            if max_width > 100:
                for cell in column:
                    current_alignment_dict = getattr(cell, 'alignment', Alignment()).__dict__
                    # 从字典中删除 wrapText，以避免重复赋值
                    current_alignment_dict.pop('wrapText', None)
                    cell.alignment = Alignment(wrapText=True, **current_alignment_dict)

    def get_sorted_merged_cells(self):
        """ 将合并单元格按照行列顺序排列。
        """
        if not hasattr(self, 'sorted_merged_cells'):
            self.sorted_merged_cells = list(sorted(self.merged_cells.ranges, key=lambda x: (x.min_row, x.min_col)))
        return self.sorted_merged_cells


inject_members(XlWorksheet, openpyxl.worksheet.worksheet.Worksheet, white_list=['_cells_by_row'])


class XlWorkbook(openpyxl.Workbook):

    def adjust_sheets(self, new_sheetnames):
        """ 按照 new_sheetnames 的清单重新调整sheets
            在清单里的按顺序罗列
            不在清单里的表格删除
            不能出现wb原本没有的表格名
        """
        for name in set(self.sheetnames) - set(new_sheetnames):
            # 最好调用标准的remove接口删除sheet
            #   不然虽然能表面上看也能删除sheet，但会有命名空间的一些冗余信息留下
            self.remove(self[name])
        self._sheets = [self[name] for name in new_sheetnames]
        return self

    def merge_sheets_by_keycol(self, sheets, keycol, new_name=None, *, cmp_func=None):
        """ 对多个工作表，按照关键列（主键）进行数据合并

        :param sheets: 多个表格（可以不同工作薄），顺序有影响，以第0个表为主表
        :param keycol: 关键字段
        :param new_name: 新的sheets名称
            todo new_name变为可选参数，不写时默认合并到第一个表格里
        :param cmp_func: 自定义匹配规则
            def cmp_func(主表已转str格式的键值, 副表已转str格式的键值):
                return True匹配成功
                return False匹配失败

        完整版实现起来有点麻烦，会循序渐进，先实现简洁版

        来自不同表的字段区分
            原本是想修改head名称来实现，比如加上前缀"表1"、"表2"，但这样有点冗余难看
            后来想到可以在每个表后面扩展一个列
                __keycol0__、__keycol1__、...
                即作为分割，也可以用于辅助计算

            todo 或者在开头加上一个合并单元格，不同表格的区分？
            todo 根据不同表格最大表头行数优化下，防止ws1表头太矮，后面有高表头的数据会复制缺失
        """
        if cmp_func is None:
            def cmp_func(k1, k2):
                return k1 == k2

        # 1 新建表格，从sheets[0]拷贝
        if new_name:
            ws1 = self.copy_worksheet(sheets[0])
            ws1.title = new_name
        else:
            ws1 = sheets[0]

        # 2 添加__keycol0__
        msg1 = ws1.find_head_data_range(keycol)
        last_right = msg1['right'] + 1
        ws1.cell(msg1['head_bottom'], last_right).value = '__keycol0__'

        exists_key = set()

        def write_new_key(row, column, value):
            ws1.cell(row, column).value = value
            if value in exists_key:
                ws1.cell(row, column).fill_color([252, 157, 154])
            else:
                exists_key.add(value)

        for i in range(msg1['data_top'], msg1['data_bottom'] + 1):
            write_new_key(i, last_right, ws1.cell(i, msg1['col']).value)

        # 3 依次把其他表添加到ws1
        last_data_bottom = msg1['data_bottom']
        for ws2_id, ws2 in enumerate(sheets[1:], start=1):
            # 3.1 ws2关键信息
            msg2 = ws2.find_head_data_range(keycol)
            data2 = []
            for i2 in range(msg2['data_top'], msg2['data_bottom'] + 1):
                data2.append([i2, str(ws2.cell(i2, msg2['col']).value)])

            # 3.2 复制表头（支持复合、合并单元格、多行格式的表头）
            msg3 = {}  # ws2复制到ws1，新区间的各种位置
            row_bias = msg2['head_bottom'] - msg1['head_bottom']  # 表头底部对齐需要的偏移行数
            msg3['head_top'] = msg2['head_top'] - row_bias
            msg3['left'] = last_right + 1
            if msg3['head_top'] < 1:  # 表头无法整个复制过来，那就要缩小下ws2表头开始的位置
                msg2['head_top'] += msg3['head_top'] + 1
                msg3['head_top'] = 1
            ws2.copy_range(excel_addr2(msg2['head_top'], msg2['left'], msg2['head_bottom'], msg2['right']),
                           ws1[excel_addr(msg3['head_top'], last_right + 1)])

            new_right = last_right + msg2['right'] - msg1['left'] + 2
            ws1.cell(msg1['head_bottom'], new_right).value = f'__keycol{ws2_id}__'

            # 3.4 先按ws1数据顺序模板填充数据
            exists_key = set()

            # trick: 配合后续f字符串的使用，把重复性的东西提前计算好了
            ws2_row_tag = excel_addr2(1, msg2['left'], 1, msg2['right']).replace('1', '{0}')

            # 考虑到可能存在重复值问题，所以这里算法是暴力双循环
            for i1 in range(msg1['data_top'], last_data_bottom + 1):
                k1 = str(ws1.cell(i1, last_right).value)
                for _i, x in enumerate(data2):
                    if cmp_func(k1, x[1]):  # todo 这里可以扩展自定义匹配规则的
                        ws2.copy_range(ws2_row_tag.format(x[0]), ws1[excel_addr(i1, msg3['left'])])
                        del data2[_i]
                        break
                else:  # ws2有，ws1没有的数据
                    pass
                write_new_key(i1, new_right, k1)

            # 3.5 剩余的data2添加到末尾
            for x in data2:
                last_data_bottom += 1
                ws2.copy_range(ws2_row_tag.format(x[0]), ws1[excel_addr(last_data_bottom, msg3['left'])])
                write_new_key(last_data_bottom, new_right, x[1])

            # 3.6 更新坐标
            last_right = new_right

    @classmethod
    def from_html(cls, content) -> 'XlWorkbook':
        from pyxllib.stdlib.tablepyxl.tablepyxl import document_to_workbook
        # 支持多 <table> 结构
        return document_to_workbook(content)

    @classmethod
    def from_latex(cls, content) -> 'XlWorkbook':
        """
        参考：kun0zhou，https://github.com/kun-zhou/latex2excel/blob/master/latex2excel.py
        """
        from openpyxl import Workbook

        # 可以处理多个表格
        wb = Workbook()
        for idx, s in enumerate(re.findall(r'(\\begin{tabular}.*?\\end{tabular})', content, flags=re.DOTALL), start=1):
            if idx == 1:
                ws = wb.active
                ws.title = 'Table 1'
            else:
                ws = wb.create_sheet(title=f'Table {idx}')
            ws.init_from_latex(s)

        return wb

    def to_html(self) -> str:
        li = []
        for ws in self.worksheets:
            li.append(ws.to_html())
        return '\n\n'.join(li)

    def to_latex(self):
        li = []
        for ws in self.worksheets:
            li.append(ws.to_latex())
        return '\n'.join(li)

    def _to_json_readonly(self, data, reduction_degree):
        if reduction_degree == 0:
            return data

        def traverse_json(obj, path=""):
            """ 递归遍历JSON对象，模拟Python字典和列表的索引机制来显示路径。
            """
            if isinstance(obj, dict):
                for name in ['filename', 'start_dir', 'header_offset', 'modified',
                             'CRC', 'compress_size', 'file_size',
                             # 230908周五16:40，后面是详细查看各种案例后，添加的处理
                             '__reduce__', '_dict', '_fonts', '_fills', '_borders',
                             '_style', '_named_styles']:
                    if name in obj:
                        del obj[name]
                for k, v in obj.items():
                    new_path = f"{path}['{k}']" if path else k
                    traverse_json(v, new_path)
            elif isinstance(obj, list):
                for i, v in enumerate(obj):
                    new_path = f"{path}[{i}]"
                    traverse_json(v, new_path)
            else:
                pass  # 对于基本数据类型，不需要进一步处理

        traverse_json(data)

        return data

    def to_json(self, reduction_degree=0):
        """
        :param reduction_degree: 对json进行处理的程度级别
            0: 最原始的json
            1: 删除易变的配置，剩下的一些属性索引使用hash值存储
            2: todo，在跨软件应用的时候，excel整体框架可能会有大改，
                此时只比较更基础性的属性，而不进行较完整的全内容比较
        """
        # 1 将对象先序列化
        s = jsonpickle.encode(self)
        data = json.loads(s)
        if reduction_degree == 0:
            return data

        if self.read_only:
            return self._to_json_readonly(data, reduction_degree)

        # 2 将复合结构hash化
        for name in ['font', 'border', 'fill']:
            ls = data[f'_{name}s']['py/seq']
            for i, x in enumerate(ls):
                ls[i] = xlmd5(json.dumps(x))

        # 3 将id引用改成其对应的hash值
        def traverse_json(obj, path=""):
            """ 递归遍历JSON对象，模拟Python字典和列表的索引机制来显示路径。
            """
            if isinstance(obj, dict):
                for k, v in obj.items():
                    for name in ['font', 'border', 'fill']:
                        if k == f'{name}Id':
                            obj[k] = data[f'_{name}s']['py/seq'][v]

                    new_path = f"{path}['{k}']" if path else k
                    traverse_json(v, new_path)
            elif isinstance(obj, list):
                for i, v in enumerate(obj):
                    new_path = f"{path}[{i}]"
                    traverse_json(v, new_path)
            else:
                pass  # 对于基本数据类型，不需要进一步处理

        traverse_json(data)

        # 4 去掉不需要对比的差异项
        def del_volatile_attrs():
            del data['properties']['modified']
            del data['properties']['created']

            del data['_fonts']  # 字体格式
            del data['_borders']  # 边框格式
            del data['_fills']  # 填充格式

            del data['_named_styles']  # 命名样式
            del data['_cell_styles']  # 单元格样式

        del_volatile_attrs()

        return data

    def to_md5(self, reduction_degree=1):
        """ 基于to_json计算的md5，一般用来判断不同workbook间是否相同 """
        return xlmd5(json.dumps(self.to_json(reduction_degree)))

    def autofit(self):
        for ws in self.worksheets:
            ws.autofit()

    def extract_summary(self, *, samples_num=5, limit_length=2500):
        """ 更新后的函数：提取整个Excel工作簿的摘要信息 """
        wb = self

        all_sheets_summary = []

        for ws in wb._sheets:  # 非数据表，也要遍历出来，所以使用了_sheets
            # 如果是标准工作表（Worksheet），使用现有的摘要提取机制
            if isinstance(ws, openpyxl.worksheet.worksheet.Worksheet):
                # 找到使用范围和表头范围
                used_range = ws.get_usedrange()
                if used_range:
                    header_range, data_range = split_header_and_data(ws, used_range)

                    # 提取表头结构
                    header_structure = extract_header_structure(ws, header_range)

                    filterRange = re.sub(r'\d+',
                                         lambda m: str(max(int(m.group()) - 1, 1)),
                                         data_range, count=1)

                    summary = ({
                        "sheetName": ws.title,
                        "sheetType": "Worksheet",
                        "usedRange": used_range,
                        "headerRange": header_range,
                        "header": header_structure,
                        'dataRange': data_range,
                        'filterRange': filterRange,
                        'sortRange': filterRange,
                        'data': extract_field_summaries(ws, header_range, data_range, samples_num)
                    })

                    if not summary['data']:  # 如果没有数据，则大概率是数据透视表，是计算出来的，读取不到~
                        summary['sheetType'] = 'PivotTable'
                        del summary['data']
                else:
                    summary = ({
                        "sheetName": ws.title,
                        "sheetType": "DialogOrMacroSheet",
                        "usedRange": None,
                    })

            # 如果是其他类型的工作表，提供基础摘要
            else:
                summary = ({
                    "sheetName": ws.title,
                    "sheetType": ws.__class__.__name__  # 使用工作表的类名作为类型
                })

            all_sheets_summary.append(summary)

        workbook_summary = {
            "fileName": Path(self.path).name if self.path else None,
            "sheetNames": wb.sheetnames,
            "sheets": all_sheets_summary,
        }

        WorkbookSummary(workbook_summary).reduce_summarys(limit_length=limit_length)

        return workbook_summary

    def extract_summary2(self):
        """ 另一套按照单元格提取摘要的程序 """
        wb = self

        all_sheets_summary = []

        for ws in wb._sheets:  # 非数据表，也要遍历出来，所以使用了_sheets
            # 如果是标准工作表（Worksheet），使用现有的摘要提取机制
            if isinstance(ws, (openpyxl.worksheet.worksheet.Worksheet)):
                # 找到使用范围和表头范围
                used_range = ws.get_usedrange()
                if used_range:
                    raw_used_range = ws.get_raw_usedrange()
                    summary = ({
                        "sheetName": ws.title,
                        "sheetType": "Worksheet",
                        "rawUsedRange": raw_used_range,
                        "usedRange": used_range,
                        'cells': extract_cells_content(ws)
                    })

                    if not summary['cells']:  # 如果没有数据，则大概率是数据透视表，是计算出来的，读取不到~
                        summary['sheetType'] = 'PivotTable'
                        del summary['cells']
                else:
                    summary = ({
                        "sheetName": ws.title,
                        "sheetType": "DialogOrMacroSheet",
                        "usedRange": None,
                    })

            # 如果是其他类型的工作表，提供基础摘要
            else:
                summary = ({
                    "sheetName": ws.title,
                    "sheetType": ws.__class__.__name__  # 使用工作表的类名作为类型
                })

            all_sheets_summary.append(summary)

        workbook_summary = {
            "fileName": Path(self.path).name if self.path else None,
            "sheetNames": wb.sheetnames,
            "sheets": all_sheets_summary,
        }

        return workbook_summary


inject_members(XlWorkbook, openpyxl.Workbook)


def __3_extract_summary():
    """ 提取表格摘要 """


def score_row(row):
    score = 0
    for cell in row:
        if cell.value is not None:
            if is_string_type(cell.value):
                score += 1  # Add positive score for string type
            else:
                score -= 1  # Subtract score for non-string type

            # 检查填充颜色和边框，为得分增加0.5分
            # if cell.fill.bgColor.rgb != '00000000' or \
            #         (cell.border.left.style or cell.border.right.style or
            #          cell.border.top.style or cell.border.bottom.style):
            #     score += 0.5
    return score


def find_header_row(ws, used_range, max_rows_to_check=10):
    """ 找到工作表中的表头行 """
    range_details = parse_range_address(used_range)

    # 初始化得分列表
    row_scores = []

    # 只检查指定的最大行数
    rows_to_check = min(range_details['bottom'] - range_details['top'] + 1, max_rows_to_check)

    # 为每行评分
    for row in ws.iter_rows(min_row=range_details['top'], max_row=range_details['top'] + rows_to_check - 1,
                            min_col=range_details['left'], max_col=range_details['right']):
        row_scores.append(score_row(row))

    # 计算行与行之间分数变化的加权
    weighted_scores = []
    for i, score in enumerate(row_scores):
        b = score - row_scores[i + 1] if i < len(row_scores) - 1 else 0
        y = score + b
        weighted_scores.append(y)

    # 确定表头行的位置
    header_row = weighted_scores.index(max(weighted_scores)) + range_details['top']

    # 从used_range的起始行到找到的表头行都视为表头
    header_range = build_range_address(left=range_details['left'], top=range_details['top'],
                                       right=range_details['right'], bottom=header_row)
    return header_range


def split_header_and_data(ws, used_range, max_rows_to_check=50):
    """ 将工作表的used_range拆分为表头范围和数据范围 """
    header_range = find_header_row(ws, used_range, max_rows_to_check)
    header_details = parse_range_address(header_range)
    used_range_details = parse_range_address(used_range)

    # 数据范围是紧接着表头下面的部分，直到used_range的结束
    data_range = build_range_address(left=used_range_details['left'], top=header_details['bottom'] + 1,
                                     right=used_range_details['right'], bottom=used_range_details['bottom'])
    return header_range, data_range


def extract_header_structure(ws, header_range):
    """ 对输入的表头位置单元格，提取表头结构 """
    header_range_details = parse_range_address(header_range)

    header_structure = {}
    merged_addresses = set()

    # 处理合并的单元格
    for merged_range in ws.merged_cells.ranges:
        # 如果合并的单元格在提供的表头范围内
        if merged_range.bounds[1] <= header_range_details['bottom'] \
                and merged_range.bounds[3] >= header_range_details['top']:
            top_left_cell = ws.cell(row=merged_range.bounds[1], column=merged_range.bounds[0])
            address = build_range_address(left=merged_range.bounds[0], top=merged_range.bounds[1],
                                          right=merged_range.bounds[2], bottom=merged_range.bounds[3])
            header_structure[address] = top_left_cell.get_render_value()
            for row in range(merged_range.bounds[1], merged_range.bounds[3] + 1):
                for col in range(merged_range.bounds[0], merged_range.bounds[2] + 1):
                    merged_addresses.add((row, col))

    # 处理未合并的单元格
    for row in ws.iter_rows(min_row=header_range_details['top'], max_row=header_range_details['bottom'],
                            min_col=header_range_details['left'], max_col=header_range_details['right']):
        for cell in row:
            # 如果这个单元格的地址还没有被添加到结构中，并且它有一个值 （后记，没有值也添加下比较好）
            if (cell.row, cell.column) not in merged_addresses:
                header_structure[cell.coordinate] = cell.get_render_value()

    return header_structure


def determine_field_type_and_summary(ws, col, start_row, end_row, rows):
    """ 根据指定的列范围确定字段的摘要信息

    :param rows: 由外部传入要抽样的数据编号
    """
    # 1 需要全量读取数据，获知主要格式，和数值范围
    data = defaultdict(list)
    for i in range(start_row, end_row + 1):
        cell = ws.cell(i, col)
        k, v = cell.get_number_format(), cell.value
        data[k].append(v)

    data2 = sorted(data.items(), key=lambda item: len(item[1]), reverse=True)
    number_formats = [x[0] for x in data2]

    # 2 获得要展示的样本值
    sample_values = []
    for i in rows:
        cell = ws.cell(i, col)
        value = cell.get_render_value()
        if isinstance(value, str) and len(value) > 20:
            value = value[:17] + '...'
        sample_values.append(value)

    # 3 数值范围（只要判断主类型的数值范围就行了）
    numeric_range = None
    for x in data2:
        try:
            fmt, values = x
            values = [v for v in values if (v is not None and not isinstance(v, str))]
            numeric_range = [min(values), max(values)]
            numeric_range[0] = xl_render_value(numeric_range[0], fmt)
            numeric_range[1] = xl_render_value(numeric_range[1], fmt)
            break
        except (TypeError, ValueError) as e:
            pass

    summary = {
        "number_formats": number_formats,
        "numeric_range": numeric_range,
        "sample_values": sample_values,
    }

    return summary


def extract_cells_content(ws, usedrange=None):
    """ 提取一个工作表中的所有单元格内容

    1、找出所有合并单元格，按照我的那个顺序先排序。记为堆栈a（sorted_merged_cells_stack）。
        另外存一个已被使用的单元格集合b（初始空）（used_cells_set）。
    2、按照A1, A2, A3,..., B1, B2,B3等顺序遍历到一个个单元格c（cell）。
    3、对于每个c，先判断是否存在b中，如果存在b中，删除b中的c，并且跳过c的处理
    4、否则判断与a的最前面的元素a0是否相交，如果相交，则从堆栈a导出a0，把a0存储到结果数组cells。并且把a0中其他衍生单元格地址存入b。
    4、如果c不与a0相交，则c自身元素值写入cells。
    """
    # 1 预备
    sorted_merged_cells_stack = ws.get_sorted_merged_cells()[::-1]
    used_cells_set = set()
    if usedrange is None:
        usedrange = ws.get_usedrange()
    usedrange_bound = parse_range_address(usedrange)
    cells = {}  # 结果集合

    # 2 遍历所有单元格
    def get_val(cell):
        val = cell.value
        if val is None:
            return ''
        else:
            return cell.get_render_value()
            # 如果感觉这步很慢，可以换一种更简洁的形式；但是发现下面这种对时间的格式显示还是太不智能。
            # if not isinstance(val, (str, int, float, bool)):
            #     return val
            # else:
            #     return str(val)

    for i in range(usedrange_bound['top'], usedrange_bound['bottom'] + 1):
        for j in range(usedrange_bound['left'], usedrange_bound['right'] + 1):
            # 2.1 合并单元格的衍生单元格，直接跳过
            if (i, j) in used_cells_set:
                used_cells_set.remove((i, j))
                continue
            cell = ws.cell(i, j)

            # 2.2 cell归属某组合并单元格（因为合并单元格排过序，所以只要很少的运算判断即可）
            if (sorted_merged_cells_stack
                    and sorted_merged_cells_stack[-1].min_row == i
                    and sorted_merged_cells_stack[-1].min_col == j):
                rng = sorted_merged_cells_stack.pop()
                for rng_i in range(rng.min_row, rng.max_row + 1):
                    for rng_j in range(rng.min_col, rng.max_col + 1):
                        used_cells_set.add((rng_i, rng_j))
                cells[rng.coord] = get_val(cell)
                used_cells_set.remove((i, j))
                continue

            # 2.3 普通单元格
            cells[cell.coordinate] = get_val(cell)

    return cells


def extract_field_summaries(ws, header_range, data_range, samples_num=5):
    """ 提取所有字段的摘要信息

    :param samples_num: 要抽样的数据行数
    """
    # 1 数据范围信息
    header_details = parse_range_address(header_range)
    data_details = parse_range_address(data_range)
    start_row = header_details['bottom'] + 1
    end_row = data_details['bottom']

    # 2 提前决定好所有字段统一抽样的数据行号
    rows = list(range(header_details['bottom'] + 1, data_details['bottom'] + 1))
    if len(rows) > samples_num:
        rows = random.sample(rows, samples_num)
        rows.sort()

    # 3 提取所有字段的摘要信息
    field_summaries = {}
    for col in ws.iter_cols(min_col=header_details['left'], max_col=header_details['right']):
        header_cell = ws.cell(header_details['bottom'], col[0].column)
        if header_cell.celltype() != 1:  # todo 这里要改成不使用衍生单元格的场景
            # 注意，原本摘要这里用的是.value，后面改成了.coordinate。原本的遇到重名就会出一些问题了~
            field_summaries[header_cell.coordinate] = determine_field_type_and_summary(
                ws, header_cell.column, start_row, end_row, rows
            )

    return field_summaries


class WorkbookSummary:
    """ 工作薄摘要相关处理功能 """

    def __init__(self, data):
        self.data = data

    def reduce_summarys(self, limit_length=2500):
        """ 精简摘要

        :param limit_length: 参考限定的长度，折算后大概是限定1500token
        """
        sheets = self.data['sheets']

        if limit_length == -1:  # -1表示不限制长度
            return sheets

        # 1 不用改变
        text1 = json.dumps(sheets, ensure_ascii=False)
        length1 = len(text1)
        if length1 < limit_length:
            return sheets

        # 2 每个字段都去掉一个样本
        for sheet in sheets:
            if 'data' in sheet:
                st_data = sheet['data']
                for col_header, col in st_data.items():
                    st_data[col_header]['sample_values'] = col['sample_values'][:-1]

        text2 = json.dumps(sheets, ensure_ascii=False)
        length2 = len(text2)
        bias = length1 - length2
        # dprint(length1, length2, bias)  # 调试用
        if length2 <= limit_length:
            return sheets

        # 3 算出理论上要去掉几个样本，才能控制到理想长度
        n = math.ceil(safe_div(length1 - limit_length, bias))
        m = 5 - n
        if m >= 0:
            for sheet in sheets:
                if 'data' in sheet:
                    st_data = sheet['data']
                    for col_header, col in st_data.items():
                        if m > 0:
                            st_data[col_header]['sample_values'] = col['sample_values'][:m]
                        elif m == 0:
                            del st_data[col_header]['sample_values']
            return sheets

        # 4 如果m<0，可能靠上述删除还不够。这应该是不可能发生的事情，但还是处理下。
        for sheet in sheets:
            if 'data' in sheet:
                del sheet['data']
            # 这个删除后可能还是太长的话，表示sheet太多了，需要删掉一些sheet

        # 如果删除所有的数据后仍然超过限制，那么删除一些表格
        while len(json.dumps(sheets, ensure_ascii=False)) > limit_length:
            sheets.pop()

        self.data['sheets'] = sheets
        return sheets

    def random_filename(self):
        self.data['fileName'] = str(random.randint(0, 2000)) + '.xlsx'

    def choice_samples(self, samples_num=5):
        """ 限定最多抽取几个样本
        """
        data = self.data
        for sheet in data['sheets']:
            if 'data' in sheet:
                # 预设好要抽哪些行
                n = min([len(v['sample_values']) for k, v in sheet['data'].items()])
                rows = list(range(n))
                if len(rows) > samples_num:
                    rows = random.sample(rows, samples_num)
                    rows.sort()
                # 抽取样本
                for col_name, col_data in sheet['data'].items():
                    col_data['sample_values'] = [col_data['sample_values'][i] for i in rows]

    def random_delete(self):
        """ 随机删除一些字段 """
        data = self.data
        for sheet in data['sheets']:
            # 80%概率删除data
            if 'data' in sheet and random.random() < 0.8:
                del sheet['data']

            # filterRange和sortRange有80%概率随机删除一个
            if 'filterRange' in sheet and 'sortRange' in sheet:
                if random.random() < 0.5:
                    del sheet['filterRange']
                else:
                    del sheet['sortRange']

            # usedRange、headRange、dataRange，有20%概率随机删掉一个
            if 'usedRange' in sheet and 'headRange' in sheet and 'dataRange' in sheet:
                r = random.random()
                if r < 1 / 3:
                    del sheet['usedRange']
                elif r < 2 / 3:
                    del sheet['headRange']
                else:
                    del sheet['dataRange']

            # hearder有50%概率随机打乱
            if random.random() < 0.5:
                sheet['header'] = shuffle_dict_keys(sheet['header'])

        # 一半的概率删掉文件名
        if random.random() < 0.5:
            del data['fileName']

        # 如果只有一个sheet，还会进一步精简
        if len(data['sheets']) == 1:
            if random.random() < 0.5:
                data = data['sheets'][0]

            for name in ['sheetName', 'sheetType']:
                if random.random() < 0.5 and name in data:
                    del data[name]

    def to_str(self):
        return json.dumps(self.data, ensure_ascii=False)


def extract_workbook_summary(file_path, mode=0,
                             samples_num=5, limit_length=2500, ignore_errors=False):
    """ 更新后的函数：提取整个Excel工作簿的摘要信息

    :param mode:
        -1，提取全量摘要（详细信息，全部样本）
        0, 标准的提取摘要（详细信息，随机抽5个样本）
        1，精简摘要，在保留逻辑完整性的前提下，随机的修改一些摘要的结构内容
    """
    try:
        wb: XlWorkbook = openpyxl.load_workbook(file_path)
    except Exception as e:
        if ignore_errors:
            return {}
        else:
            raise e

    if mode == -1:
        res = wb.extract_summary(samples_num=1000, limit_length=-1)
        res['fileName'] = Path(file_path).name
    elif mode == 0:
        res = wb.extract_summary(samples_num=samples_num, limit_length=limit_length)
        res['fileName'] = Path(file_path).name

    elif mode == 1:
        res = wb.extract_summary(samples_num=samples_num)

        wb_summary = WorkbookSummary(res)
        wb_summary.random_filename()
        wb_summary.random_delete()
        wb_summary.reduce_summarys(limit_length=limit_length)

        res = wb_summary.data
    else:
        raise ValueError('mode参数值不正确')

    return res


def extract_workbook_summary2(file_path, *,
                              keep_links=False,
                              keep_vba=False,
                              mode=0,
                              return_mode=0,
                              **kwargs):
    """
    :param keep_links: 是否保留外部表格链接数据。如果保留，打开好像会有点问题。
    :param mode:
        0，最原始的summary3摘要
        1，添加当前工作表、单元格位置的信息
    :param kwargs: 捕捉其他参数，主要是向下兼容，其实现在并没有用

    注意这里没有提供read_only可选参数，是因为read_only=True模式在我这里是运行不了的。
    """

    # 1 读取文件wb
    file_path = Path(file_path)
    res = {}
    res['fileName'] = file_path.name
    start_time = time.time()
    wb = load_as_xlsx_file(file_path, keep_links=keep_links, keep_vba=keep_vba)
    load_time = time.time() - start_time
    if wb is None:  # 不支持的文件类型，不报错，只是返回最基本的文件名信息
        if return_mode == 1:
            return res, load_time
        else:
            return res

    # 2 提取摘要
    summary2 = wb.extract_summary2()
    DictTool.ior(res, summary2)
    if mode == 1:
        ws = wb.active
        res['ActiveSheet'] = ws.title
        if hasattr(ws, 'selected_cell'):
            res['Selection'] = ws.selected_cell

    # res = convert_to_json_compatible(res)

    if return_mode == 1:
        return res, load_time
    else:
        return res


def update_raw_summary2(data):
    # 1 中文率
    if 'chineseContentRatio' not in data:
        texts = [data['fileName']]  # 文件名和表格名都要加上
        texts += [x for x in data['sheetNames']]

        texts += [v for sheet in data['sheets'] for v in sheet.get('cells', {}).values() if v]
        all_text = ''.join(map(str, texts))
        data['chineseContentRatio'] = round(calc_chinese_ratio(all_text), 4)

    # 2 非空单元格率
    if 'nonEmptyCellRatio' not in data:
        content_area, total_area = 0, 0
        for sheet in data['sheets']:
            for addr, value in sheet.get('cells', {}).items():
                area = get_addr_area(addr)
                if value != '':
                    content_area += area
                total_area += area
        data['nonEmptyCellRatio'] = round(safe_div(content_area, total_area), 4)

    # 3 判断键值顺序
    keys = list(data.keys())
    ref_keys = ['fileName', 'chineseContentRatio', 'nonEmptyCellRatio', 'sheetNames', 'sheets']
    if keys != ref_keys:
        data = {k: data[k] for k in ref_keys if k in data}

    return data


def extract_workbook_summary2plus(file_path, **kwargs):
    """ 增加了全局ratio的计算 """
    # 1 主体摘要
    data = extract_workbook_summary2(file_path, **kwargs)
    if not data:
        return data

    # 2 增加一些特征计算
    # todo 后续估计要改成按table的颗粒度统计以下特征

    data = update_raw_summary2(data)
    return data


class WorkbookSummary3:
    """ 计算summary3及衍生版本需要的一些功能组件 """

    @classmethod
    def count_length(cls, text):
        if not isinstance(text, str):
            text = json.dumps(text, ensure_ascii=False, default=str)
        return len(text)

    @classmethod
    def reduce1_delete_empty_cell(cls, summary3):
        """ 删除空单元格 """
        for sheet in summary3['sheets']:
            new_cells = {}
            for addr, val in sheet['cells'].items():
                if val != '':
                    new_cells[addr] = val
            sheet['cells'] = new_cells

    @classmethod
    def reduce2_truncate_overlong_cells(cls, summary3, summary_limit_len, *, cur_summary_len=None):
        """ 截断过长的单元格内容以满足表格摘要总长度限制。

        此算法旨在处理单个或多个单元格内容过长时，整个表格摘要总长度超过预设限制的问题。在保留尽可能多的有用信息的同时，智能地缩减内容长度。

        :param dict summary3: 表格摘要数据，包含多个sheet及其单元格内容
        :param int summary_limit_len: 表格摘要的最大长度限制
        :param int cur_summary_len: 当前表格摘要的长度，如果不提供，则会计算
        :return int: 调整后的表格摘要长度

        算法执行流程：
        1. 计算基准单元格长度和当前摘要长度超出部分（delta_length）。
        2. 筛选出超过基准长度的单元格并按长度降序排序。
        3. 逐个尝试截断单元格直到总长度满足要求或处理完所有单元格。
        """

        # 如果未提供当前摘要长度，则计算之
        if cur_summary_len is None:
            cur_summary_len = cls.count_length(summary3)

        # 1. 计算基准单元格长度
        total_cells_num = sum(len(st['cells']) + 5 for st in summary3['sheets'])
        base_cell_length = int(-60 * math.log(total_cells_num, 10) + 260)  # 从200趋近到20
        base_cell_length = min(int(summary_limit_len * 0.05), base_cell_length)
        base_cell_length = max(int(summary_limit_len * 0.005), base_cell_length)

        # 2. 预提取全部单元格数据信息并计算需要减少的长度
        delta_length = cur_summary_len - summary_limit_len  # 计算需要减少的长度
        overlong_cells = [(sheet, addr, val, len(val)) for sheet in summary3['sheets']
                          for addr, val in sheet['cells'].items() if
                          isinstance(val, str) and len(val) > base_cell_length]
        overlong_cells.sort(key=lambda x: -x[3])  # 按长度降序排序

        # 3. 逐个尝试截断单元格直到满足长度要求
        possible_reduction = 0
        for i, (_, _, _, length) in enumerate(overlong_cells):
            next_len = overlong_cells[i + 1][3] if i + 1 < len(overlong_cells) else base_cell_length
            possible_reduction += (length - next_len) * (i + 1)
            if possible_reduction >= delta_length or i == len(overlong_cells) - 1:
                for j in range(i + 1):
                    sheet, addr, val, _ = overlong_cells[j]
                    sheet['cells'][addr] = val[:next_len - 3] + '...'  # 更新单元格内容
                break

        return cur_summary_len - possible_reduction

    @classmethod
    def reduce3_fold_rows(cls, summary3, summary_limit_len, *, cur_summary_len=None):
        if cur_summary_len is None:
            cur_summary_len = cls.count_length(summary3)

        # 每个sheet本身其他摘要，按照5个单元格估算
        total_cells_num = sum([(len(st['cells']) + 5) for st in summary3['sheets']])
        avg_cell_len = cur_summary_len / total_cells_num
        # 目标删除单元格数量，向上取整
        target_reduce_cells_num = int((cur_summary_len - summary_limit_len) / avg_cell_len + 0.5)

        # 相同区域，头尾至少要留2行，然后考虑压缩量，一般一块range至少要10行同构数据，进行中间至少6行的压缩描述才有意义
        # 考虑重要性，应该是从末尾表格，末尾数据往前检索压缩，直到压缩量满足要求

        for sheet in reversed(summary3['sheets']):
            cells = sheet['cells']
            # 1 对单元格，先按行分组
            last_line_id = -1
            row_groups = []
            for addr, val in cells.items():
                m = re.search(r'\d+', addr)
                if not m:  # 应该都一定能找到的，这个判断是为了防止报错
                    continue
                line_id = int(m.group())

                val_type_tag = '@' if isinstance(val, str) else '#'
                cell_tag = re.sub(r'\d+', '', addr) + val_type_tag

                if line_id == last_line_id:
                    row_groups[-1].append([addr, cell_tag])
                else:
                    row_groups.append([[addr, cell_tag]])
                last_line_id = line_id

            # 2 算出每一行的row_tag，并按照row_tag再分组
            last_row_tag = ''
            rows_groups = []
            for row in row_groups:
                row_tag = ''.join([cell_tag for _, cell_tag in row])
                if row_tag == last_row_tag:
                    rows_groups[-1].append(row)
                else:
                    rows_groups.append([row])
                last_row_tag = row_tag

            # 3 开始压缩
            def extract_cells_from_rows(rows):
                for row in rows:
                    for addr, _ in row:
                        new_cells[addr] = cells[addr]

            new_cells = {}
            for rows in rows_groups:
                if len(rows) < 10:
                    extract_cells_from_rows(rows)
                else:  # 压缩中间的数据
                    # 如果评估到最终摘要可能太小，要收敛下删除的范围
                    n, m = len(rows), len(rows[0])
                    target_n = int(target_reduce_cells_num / m + 0.5)  # 本来应该删除多少行才行
                    cur_n = n - 4 if target_n > n - 4 else target_n  # 实际删除多少行
                    left_n = n - cur_n  # 剩余多少行
                    b = left_n // 2
                    a = left_n - b

                    extract_cells_from_rows(rows[:a])
                    addr = combine_addresses(rows[a][0][0], rows[-b - 1][-1][0])
                    # new_cells[addr] = '这块区域的内容跟前面几行、后面几行的内容结构是一致的，省略显示'
                    new_cells[addr] = '...'
                    extract_cells_from_rows(rows[-b:])

                    target_reduce_cells_num -= cur_n * m
                    if target_reduce_cells_num <= 0:
                        break

            sheet['cells'] = new_cells

    @classmethod
    def reduce4_truncate_cells(cls, y, summary_limit_len, *, cur_summary_len=None):
        if cur_summary_len is None:
            cur_summary_len = cls.count_length(y)

        # 1 预计要删除单元格数
        sheet_cells_num = [len(st['cells']) for st in y['sheets']]
        # 每个sheet本身其他摘要，按照5个单元格估算
        total_cells_num = sum(sheet_cells_num) + len(sheet_cells_num) * 5
        avg_cell_len = cur_summary_len / total_cells_num
        # 目标删除单元格数量，向上取整
        target_reduce_cells_num = int((cur_summary_len - summary_limit_len) / avg_cell_len + 0.5)

        # 2 所有的单元格如果都不够删，那就先把所有cells删了再说
        if total_cells_num < target_reduce_cells_num:
            for st in y['sheets']:
                st['cells'] = {}
            return cls.count_length(y)

        # 3 否则每张表按照比例删单元格，只保留前面部分的单元格
        left_rate = 1 - target_reduce_cells_num / total_cells_num
        while True:
            for i, st in enumerate(y['sheets']):
                st['cells'] = dict(islice(st['cells'].items(), int(left_rate * sheet_cells_num[i])))
            cur_summary_len = cls.count_length(y)
            if cur_summary_len <= summary_limit_len:
                return cur_summary_len
            if left_rate * total_cells_num < 1:
                break
            else:  # 缩小保留比例，再试
                left_rate *= 0.8

        return cur_summary_len

    @classmethod
    def reduce5_truncate_sheets(cls, y, summary_limit_len, *, cur_summary_len=None):
        """ 计算平均每张表的长度，保留前面部分的表格 """
        if cur_summary_len is None:
            cur_summary_len = cls.count_length(y)

        n = len(y['sheets'])
        avg_sheet_len = cur_summary_len / n
        target_reduce_sheet_num = int((cur_summary_len - summary_limit_len) / avg_sheet_len + 0.5)
        y['sheets'] = y['sheets'][:n - target_reduce_sheet_num]

        while y['sheets']:
            cur_summary_len = cls.count_length(y)
            if cur_summary_len <= summary_limit_len:
                return cur_summary_len
            y['sheets'] = y['sheets'][:-1]  # 依次尝试删除最后一张表格的详细信息

    @classmethod
    def summary2_to_summary3(cls, summary2, summary_limit_len=4000):
        """ 将summary2转换成summary3 """

        def reduce_step_by_step(y):
            mode_tags = [
                'Delete empty cell',
                'Omit the longer content and replace it with...',
                'Omit lines with the same structure',
                'Omit later lines',
                'Omit later sheets'
            ]

            # 0 摘要本来就不大
            cur_summary_len = cls.count_length(y)
            if cur_summary_len <= summary_limit_len:
                return y

            # 1 删除空单元格
            cls.reduce1_delete_empty_cell(y)
            cur_summary_len = cls.count_length(y)
            if cur_summary_len <= summary_limit_len:
                y['mode'] = ', '.join(mode_tags[:1])
                return y

            # 2 单个单元格内容过长的，省略显示
            cur_summary_len = cls.reduce2_truncate_overlong_cells(y, summary_limit_len, cur_summary_len=cur_summary_len)
            if cur_summary_len <= summary_limit_len:
                y['mode'] = ', '.join(mode_tags[:2])
                return y

            # 3 同构数据，省略显示（有大量相同行数据，折叠省略表达）
            cls.reduce3_fold_rows(y, summary_limit_len, cur_summary_len=cur_summary_len)
            cur_summary_len = cls.count_length(y)
            if cur_summary_len <= summary_limit_len:
                y['mode'] = ', '.join(mode_tags[:3])
                return y

            # 4 每张表都按比例删除后面部分的单元格
            cur_summary_len = cls.reduce4_truncate_cells(y, summary_limit_len, cur_summary_len=cur_summary_len)
            if cur_summary_len <= summary_limit_len:
                y['mode'] = ', '.join(mode_tags[:4])
                return y

            # 5 从后往前删每张表格的详细信息
            cls.reduce5_truncate_sheets(y, summary_limit_len, cur_summary_len=cur_summary_len)
            y['mode'] = ', '.join(mode_tags[:5])
            return y

        x = summary2
        y = {
            'fileName': x['fileName'],
            'sheetNames': x['sheetNames'],
            'sheets': x['sheets'],
            'mode': 'Complete information',
        }

        # 处理前确保下cells字段存在，避免后续很多处理过程要特判
        for st in y['sheets']:
            if 'cells' not in st:
                st['cells'] = {}

        y = reduce_step_by_step(y)

        # 但是最后结果还是去掉空cells
        for st in y['sheets']:
            if not st['cells']:
                del st['cells']

        return y

    @classmethod
    def reduce4b(cls, y, summary_limit_len, *, cur_summary_len=None, active_sheet_weight=0.5):
        """
        :param active_sheet_weight: 当前活动表格被删除的权重，0.5表示按比例被删除的量只有其他表格的一半
        """
        if cur_summary_len is None:
            cur_summary_len = cls.count_length(y)

        active_sheet = y['ActiveSheet']

        # 1 预计要删除单元格数
        sheet_cells_num = [len(st['cells']) for st in y['sheets']]
        # 每个sheet本身其他摘要，按照5个单元格估算
        total_cells_num = sum(sheet_cells_num) + len(sheet_cells_num) * 5
        avg_cell_len = cur_summary_len / total_cells_num
        # 目标删除单元格数量，向上取整
        target_reduce_cells_num = int((cur_summary_len - summary_limit_len) / avg_cell_len + 0.5)

        # 2 对当前活动表格，会减小删除权重
        # 标记当前活动表格的单元格数
        active_sheet_index = [i for i, st in enumerate(y['sheets']) if st['sheetName'] == active_sheet][0]
        active_cells_num = sheet_cells_num[active_sheet_index]

        # 计算权重系数
        w = active_sheet_weight  # 当前激活表的权重系数
        m = active_cells_num
        n = total_cells_num
        r = target_reduce_cells_num / n

        # 计算非活动表格的额外权重系数
        w2 = 1 + m * (1 - w) / (n - m)

        # 3 所有的单元格如果都不够删，那就先把所有cells删了再说
        if total_cells_num < target_reduce_cells_num:
            for st in y['sheets']:
                st['cells'] = {}
            return cls.count_length(y)

        # 4 否则每张表按照比例删单元格，只保留前面部分的单元格
        left_rate = 1 - r  # 原始保留比例
        while True:
            for i, st in enumerate(y['sheets']):
                if i == active_sheet_index:
                    # 当前激活的sheet保留更多单元格
                    st['cells'] = dict(islice(st['cells'].items(), int(left_rate * w * sheet_cells_num[i])))
                else:
                    # 其他sheet按照w2权重删除单元格
                    st['cells'] = dict(islice(st['cells'].items(), int(left_rate * w2 * sheet_cells_num[i])))
            cur_summary_len = cls.count_length(y)
            if cur_summary_len <= summary_limit_len:
                return cur_summary_len
            if left_rate * total_cells_num < 1:
                break
            else:
                left_rate *= 0.8  # 缩小保留比例，再试

        return cur_summary_len

    @classmethod
    def reduce5b(cls, y, summary_limit_len, *, cur_summary_len=None):
        """ 计算平均每张表的长度，保留前面部分的表格 """
        if cur_summary_len is None:
            cur_summary_len = cls.count_length(y)

        n = len(y['sheets'])
        active_sheet_name = y['ActiveSheet']

        avg_sheet_len = cur_summary_len / n
        # target_reduce_sheet_num = int((cur_summary_len - summary_limit_len) / avg_sheet_len + 0.5)
        # y['sheets'] = y['sheets'][:n - target_reduce_sheet_num]

        while y['sheets']:
            cur_summary_len = cls.count_length(y)
            if cur_summary_len <= summary_limit_len:
                return cur_summary_len

            # 如果最后一张表格是激活的表格，尝试删除前一张
            if y['sheets'][-1]['sheetName'] == active_sheet_name:
                if len(y['sheets']) > 1:
                    y['sheets'] = y['sheets'][:-2] + [y['sheets'][-1]]
                else:
                    y['sheets'] = []
            else:
                y['sheets'] = y['sheets'][:-1]  # 删除最后一张表格的详细信息

        return cur_summary_len

    @classmethod
    def summary2_to_summary3b(cls, summary2, summary_limit_len=4000):
        """ 将summary2转换成summary3 """

        def reduce_step_by_step(y):
            mode_tags = [
                'Delete empty cell',
                'Omit the longer content and replace it with...',
                'Omit lines with the same structure',
                'Omit later lines',
                'Omit later sheets'
            ]

            # 0 摘要本来就不大
            cur_summary_len = cls.count_length(y)
            if cur_summary_len <= summary_limit_len:
                return y

            # 1 删除空单元格
            cls.reduce1_delete_empty_cell(y)
            cur_summary_len = cls.count_length(y)
            if cur_summary_len <= summary_limit_len:
                y['mode'] = ', '.join(mode_tags[:1])
                return y

            # 2 单个单元格内容过长的，省略显示
            cur_summary_len = cls.reduce2_truncate_overlong_cells(y, summary_limit_len, cur_summary_len=cur_summary_len)
            if cur_summary_len <= summary_limit_len:
                y['mode'] = ', '.join(mode_tags[:2])
                return y

            # 3 同构数据，省略显示（有大量相同行数据，折叠省略表达）
            cls.reduce3_fold_rows(y, summary_limit_len, cur_summary_len=cur_summary_len)
            cur_summary_len = cls.count_length(y)
            if cur_summary_len <= summary_limit_len:
                y['mode'] = ', '.join(mode_tags[:3])
                return y

            # 4 每张表都按比例删除后面部分的单元格
            cur_summary_len = cls.reduce4b(y, summary_limit_len, cur_summary_len=cur_summary_len)
            if cur_summary_len <= summary_limit_len:
                y['mode'] = ', '.join(mode_tags[:4])
                return y

            # 5 从后往前删每张表格的详细信息
            cls.reduce5b(y, summary_limit_len, cur_summary_len=cur_summary_len)
            y['mode'] = ', '.join(mode_tags[:5])
            return y

        x = summary2
        y = {
            'fileName': x['fileName'],
            'sheetNames': x['sheetNames'],
            'sheets': x['sheets'],
            'mode': 'Complete information',
            'ActiveSheet': x['ActiveSheet'],  # 当期激活的工作表
        }
        if 'Selection' in x:
            # 最多截取250个字符。（一般情况下这个很小的，只是在很极端情况，比如离散选中了非常多区域，这个可能就会太长
            y['Selection'] = x['Selection'][:250]

        # 处理前确保下cells字段存在，避免后续很多处理过程要特判
        for st in y['sheets']:
            if 'cells' not in st:
                st['cells'] = {}

        y = reduce_step_by_step(y)

        # 但是最后结果还是去掉空cells
        for st in y['sheets']:
            if not st['cells']:
                del st['cells']

        return y


class WorkbookSummary3plus(WorkbookSummary3):
    """ 标准的token计算方式，不过暂不打算实装使用 """

    @classmethod
    def count_length(cls, text):
        from pyxlpr.data.gptlib import Tokenizer
        if not isinstance(text, str):
            text = json.dumps(text, ensure_ascii=False, default=str)
        return Tokenizer.count_tokens(text)


def extract_workbook_summary3(file_path, summary_limit_len=4000, **kwargs):
    """ 增加了全局ratio的计算 """
    data = extract_workbook_summary2(file_path, **kwargs)
    if not data:
        return data
    data = WorkbookSummary3.summary2_to_summary3(data, summary_limit_len)
    return data


def extract_workbook_summary3b(file_path,
                               summary_limit_len=4000,
                               timeout_seconds=10,
                               return_mode=0,
                               debug=False,
                               len_mode=0,
                               **kwargs):
    """

    :param summary_limit_len: 摘要长度限制
    :param timeout_seconds: 超时限制
    :param return_mode: 返回模式，0表示只返回摘要，1表示返回摘要和耗时
    :param len_mode:
        0, 使用len作为token长度评估
        1, 使用模型评估实际token长度
    :param kwargs: 其他是summary2读取文件的时候的参数，其实都不太关键，一般不用特地设置
    """
    res = {}
    res['fileName'] = Path(file_path).name
    load_time = summary2_time = summary3_time = -1

    try:
        with Timeout(timeout_seconds):
            start_time = time.time()
            res, load_time = extract_workbook_summary2(file_path, mode=1, return_mode=1, **kwargs)
            # res = convert_to_json_compatible(res)
            summary2_time = time.time() - start_time - load_time
            start_time = time.time()
            if len_mode == 1:
                res = WorkbookSummary3plus.summary2_to_summary3b(res, summary_limit_len)
            else:
                res = WorkbookSummary3.summary2_to_summary3b(res, summary_limit_len)
            summary3_time = time.time() - start_time
    except TimeoutError as e:
        if debug:
            raise e
        res['error'] = f'超时，未完成摘要提取：{timeout_seconds}秒'
    except Exception as e:
        if debug:
            raise e
        res['error'] = f'提取摘要时发生错误：{format_exception(e, 2)}'

    if return_mode == 1:
        return res, {'load_time': human_readable_number(load_time),
                     'summary2_time': human_readable_number(summary2_time),
                     'summary3_time': human_readable_number(summary3_time)}
    return res
